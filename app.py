import io
import re
import json
import base64
import pytz
import streamlit as st
import pandas as pd
from google import genai
from google.genai import types
from supabase import create_client, Client
from datetime import datetime

# ─────────────────────────────────────────
# 1. CONFIGURATION
# ─────────────────────────────────────────
try:
    SUPABASE_URL  = st.secrets["SUPABASE_URL"]
    SUPABASE_KEY  = st.secrets["SUPABASE_KEY"]
    GEMINI_KEY    = st.secrets["GENAI_API_KEY"]
    TBL_BOOKINGS          = "bookings"
    TBL_REVISIONS         = "booking_revisions"
    TBL_LOCAL_CHARGES     = "local_charges"
    TBL_LOCAL_CHARGES_V2  = "local_charges_v2"
    TBL_LOCAL_CHARGE_ITEMS = "local_charge_items"
except Exception:
    # สำหรับใช้รันในเครื่องตัวเอง (Local) ถ้ายังไม่ได้ตั้งค่า Secrets
    st.error("❌ ไม่พบ API Keys ในระบบ Secrets กรุณาตั้งค่าที่ Settings > Secrets")
    st.stop()
    TBL_BOOKINGS           = "test_bookings"
    TBL_REVISIONS          = "test_booking_revisions"
    TBL_LOCAL_CHARGES      = "test_local_charges"
    TBL_LOCAL_CHARGES_V2   = "test_local_charges_v2"
    TBL_LOCAL_CHARGE_ITEMS = "test_local_charge_items"

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
genai_client     = genai.Client(api_key=GEMINI_KEY)
GEMINI_MODEL_BOOKING = "models/gemini-2.5-flash"
GEMINI_MODEL         = "models/gemini-3.1-flash-lite-preview"

# ─────────────────────────────────────────
# 2. PAGE CONFIG
# ─────────────────────────────────────────
st.set_page_config(
    page_title="DHL Booking Cloud Extractor",
    page_icon="🚚",
    layout="wide",
)
 
# ─────────────────────────────────────────
# 3. CSS THEME  (DHL — Golden gradient style)
# ─────────────────────────────────────────
st.markdown("""
<style>
    .stApp { background: linear-gradient(135deg, #FFCC00 0%, #FFD700 50%, #ba9500 100%); }
    .block-container { background-color: white; padding: 40px; border-radius: 25px; box-shadow: 0 15px 35px rgba(0,0,0,0.3); border: 6px solid #D40511; margin-top: 20px; margin-bottom: 20px; }
    .stDataFrame, div[data-testid="stTable"] { background-color: #f0f2f6 !important; border-radius: 10px; padding: 10px; box-shadow: inset 2px 2px 5px rgba(0,0,0,0.05); }
    .stButton>button { background-color: #D40511; color: white; border-radius: 10px; border: none; box-shadow: 0 4px #990000; transition: 0.2s; width: 100%; }
    .stButton>button:hover { background-color: #ff0000; transform: translateY(-2px); box-shadow: 0 6px #990000; }
    [data-testid="stDownloadButton"] > button { background-color: #D40511; color: white; border-radius: 10px; border: none; box-shadow: 0 4px #990000; transition: 0.2s; width: auto; }
    [data-testid="stDownloadButton"] > button:hover { background-color: #ff0000; transform: translateY(-2px); box-shadow: 0 6px #990000; }
    [data-testid="metric-container"] { background: #fff9e6; border: 2px solid #FFCC00; border-radius: 12px; padding: 16px 20px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); }
    [data-testid="metric-container"] [data-testid="stMetricValue"] { color: #D40511; font-size: 28px; font-weight: 700; }
    [data-testid="stFileUploader"] { border: 2px dashed #FFCC00 !important; border-radius: 12px !important; padding: 8px !important; background-color: #fffef5 !important; }
    [data-testid="stFileUploader"]:hover { border-color: #D4A900 !important; background-color: #fffbe6 !important; }
    [data-testid="stTextInput"] > div { border: 2px solid #FFCC00 !important; border-radius: 10px !important; background-color: #fffef5 !important; }
    [data-testid="stTextInput"] > div:focus-within { border-color: #D4A900 !important; box-shadow: 0 0 0 3px rgba(255,204,0,0.2) !important; }
    [data-testid="stSelectbox"] > div > div { border: 2px solid #FFCC00 !important; border-radius: 10px !important; background-color: #fffef5 !important; }
    [data-testid="stSelectbox"] > div > div:focus-within { border-color: #D4A900 !important; box-shadow: 0 0 0 3px rgba(255,204,0,0.2) !important; }
    [data-testid="stTextArea"] > div { border: 2px solid #FFCC00 !important; border-radius: 10px !important; background-color: #fffef5 !important; }
    [data-testid="stTextArea"] > div:focus-within { border-color: #D4A900 !important; box-shadow: 0 0 0 3px rgba(255,204,0,0.2) !important; }
</style>
""", unsafe_allow_html=True)
 
# ─────────────────────────────────────────
# 4. HEADER
# ─────────────────────────────────────────
_truck_svg = """<svg viewBox="0 0 230 92" xmlns="http://www.w3.org/2000/svg" width="230" height="92">
  <ellipse cx="115" cy="90" rx="108" ry="4" fill="rgba(0,0,0,0.07)"/>
  <rect x="78" y="71" width="146" height="6" rx="1" fill="#444" stroke="#222" stroke-width="1"/>
  <rect x="74" y="6" width="150" height="65" rx="3" fill="#FFCC00" stroke="#222" stroke-width="2.5"/>
  <rect x="74" y="6" width="150" height="7" rx="2" fill="#D4A900" stroke="#222" stroke-width="1.5"/>
  <rect x="74" y="64" width="150" height="7" rx="1" fill="#D4A900" stroke="#222" stroke-width="1.5"/>
  <line x1="112" y1="13" x2="112" y2="64" stroke="#D4A900" stroke-width="2"/>
  <line x1="150" y1="13" x2="150" y2="64" stroke="#D4A900" stroke-width="2"/>
  <line x1="188" y1="13" x2="188" y2="64" stroke="#D4A900" stroke-width="2"/>
  <rect x="218" y="13" width="6" height="51" rx="1" fill="#D4A900" stroke="#222" stroke-width="1"/>
  <text x="149" y="51" font-family="Arial Black,Arial,sans-serif" font-size="27" font-weight="900" fill="#D40511" text-anchor="middle" letter-spacing="-1">DHL</text>
  <rect x="4" y="10" width="66" height="63" rx="5" fill="#D40511" stroke="#222" stroke-width="2.5"/>
  <rect x="4" y="10" width="13" height="63" rx="4" fill="#B8030F" stroke="#222" stroke-width="1"/>
  <rect x="18" y="15" width="36" height="27" rx="3" fill="#AED6F1" stroke="#333" stroke-width="1.5"/>
  <line x1="22" y1="18" x2="27" y2="37" stroke="white" stroke-width="2" opacity="0.4"/>
  <rect x="54" y="15" width="12" height="18" rx="2" fill="#AED6F1" stroke="#333" stroke-width="1"/>
  <line x1="52" y1="43" x2="52" y2="71" stroke="#aa0000" stroke-width="1.5"/>
  <rect x="54" y="53" width="10" height="3" rx="1.5" fill="#FFCC00" stroke="#D4A900" stroke-width="1"/>
  <rect x="5" y="18" width="10" height="6" rx="2" fill="#FFF9C4" stroke="#333" stroke-width="1"/>
  <rect x="5" y="50" width="10" height="5" rx="2" fill="#FF8A65" stroke="#333" stroke-width="1"/>
  <rect x="5" y="26" width="10" height="22" rx="1" fill="#222" stroke="#111" stroke-width="1"/>
  <line x1="5" y1="31" x2="15" y2="31" stroke="#555" stroke-width="1"/>
  <line x1="5" y1="36" x2="15" y2="36" stroke="#555" stroke-width="1"/>
  <line x1="5" y1="41" x2="15" y2="41" stroke="#555" stroke-width="1"/>
  <rect x="4" y="67" width="66" height="6" rx="2" fill="#999" stroke="#333" stroke-width="1.5"/>
  <rect x="-1" y="18" width="8" height="6" rx="1" fill="#777" stroke="#444" stroke-width="1"/>
  <line x1="4" y1="21" x2="7" y2="21" stroke="#555" stroke-width="1.5"/>
  <rect x="60" y="0" width="5" height="16" rx="2" fill="#888" stroke="#555" stroke-width="1"/>
  <circle cx="63" cy="-1" r="3" fill="#bbb" opacity="0.4"/>
  <circle cx="65" cy="-5" r="2" fill="#bbb" opacity="0.25"/>
  <rect x="62" y="67" width="16" height="5" rx="1" fill="#666" stroke="#333" stroke-width="1"/>
  <circle cx="19" cy="80" r="11" fill="#1a1a1a" stroke="#555" stroke-width="2"/>
  <circle cx="19" cy="80" r="5.5" fill="#666"/>
  <circle cx="19" cy="80" r="2" fill="#333"/>
  <circle cx="52" cy="80" r="11" fill="#1a1a1a" stroke="#555" stroke-width="2"/>
  <circle cx="52" cy="80" r="5.5" fill="#666"/>
  <circle cx="52" cy="80" r="2" fill="#333"/>
  <circle cx="67" cy="80" r="11" fill="#1a1a1a" stroke="#555" stroke-width="2"/>
  <circle cx="67" cy="80" r="5.5" fill="#666"/>
  <circle cx="67" cy="80" r="2" fill="#333"/>
  <circle cx="168" cy="80" r="11" fill="#1a1a1a" stroke="#555" stroke-width="2"/>
  <circle cx="168" cy="80" r="5.5" fill="#666"/>
  <circle cx="168" cy="80" r="2" fill="#333"/>
  <circle cx="207" cy="80" r="11" fill="#1a1a1a" stroke="#555" stroke-width="2"/>
  <circle cx="207" cy="80" r="5.5" fill="#666"/>
  <circle cx="207" cy="80" r="2" fill="#333"/>
</svg>"""

_truck_b64 = base64.b64encode(_truck_svg.encode()).decode()

st.markdown(f"""
    <div style="display: flex; align-items: center; margin-bottom: 20px; padding: 15px; background-color: #ffffff; border-radius: 12px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); border-left: 10px solid #FFCC00;">
        <div style="margin-right: 20px; flex-shrink: 0;">
            <img src="data:image/svg+xml;base64,{_truck_b64}" width="230" height="92"/>
        </div>
        <div style="flex-grow: 1;">
            <h1 style="margin: 0; color: #333; font-size: 30px; line-height: 1.2;">DSC: CTC FG Export</h1>
            <p style="margin: 0; color: #666; font-size: 18px;">Booking Cloud Extractor (Vision Engine)</p>
        </div>
    </div>
""", unsafe_allow_html=True)
 
# ─────────────────────────────────────────
# 5. HELPERS
# ─────────────────────────────────────────
COLUMNS_ORDER = [
    "booking_no", "loading_at", "fcl_or_lcl", "by_air_or_sea",
    "country", "port_of_destination", "liner_name", "vessel_name",
    "no_container", "container_type", "no_pallet",
    "etd", "eta", "liner_cutoff", "vgm_cutoff", "si_cutoff",
    "cy_date", "cy_at", "return_date_1st", "return_place",
    "paperless_code", "updated_at",
]
 
PROMPT_BOOKING = """You are a DHL Logistics Analyst. Extract ALL shipping info from this booking PDF.

Return ONLY a JSON object (no markdown, no explanation):
{
  "booking_no":          "Carrier Ref or Booking No.",
  "fcl_or_lcl":          "FCL or LCL",
  "by_air_or_sea":       "Air or Sea",
  "country":             "destination country (NOT Thailand)",
  "port_of_destination": "final destination of the shipment",
  "liner_name":          "shipping line",
  "vessel_name":         "vessel/voyage (include connecting if any)",
  "no_container":        number or null,
  "container_type":      "ALL container counts+types combined e.g. '1X40HC+1X20GP' or '2X40HC' — null if LCL or no container",
  "no_pallet":           number or null,
  "cy_at":               "empty pick-up depot",
  "return_place":        "laden return location",
  "paperless_code":      "4-digit code from PAPERLESS CODE or PORT CODE label e.g. 2836, or null",
  "liner_cutoff":        "dd/mm/yyyy hh:mm or null",
  "vgm_cutoff":          "dd/mm/yyyy hh:mm or null",
  "si_cutoff":           "dd/mm/yyyy hh:mm or null",
  "return_date_1st":     "dd/mm/yyyy or null",
  "cy_date":             "dd/mm/yyyy or null",
  "etd":                 "dd/mm/yyyy or null",
  "eta":                 "dd/mm/yyyy or null"
}

Rules:
- booking_no: extract using this priority order:
  1. Value next to labels "Carrier Booking No.", "Carrier Booking Reference", "Carrier Ref" — use that directly.
  2. Value next to label "Booking No." or "Booking No" — BUT only if the value looks like an actual booking number (alphanumeric code). If the value is descriptive text (e.g. "LOAD ON SHIPPER NAME", "SEE ATTACHED", or any phrase that is clearly not a code), skip it.
  3. Value next to "Ref No", "Ref No.", "Reference No." — use as fallback if steps 1 and 2 yield nothing.
  4. Return null if nothing found.
  Do NOT use B/L No., forwarder ref (e.g. FLXCB-), CONSOL, or tracking number.
- port_of_destination: final destination of the shipment — use the "To:" field in the booking header, or the last stop in the Intended Transport Plan. Do NOT use intermediate sea ports or terminal names (e.g. "Guadalajara Castilla La Mancha, Spain" not "APM Terminal Valencia").
- country: use consignee's country if clearly stated in address; otherwise infer from port_of_destination.
- cy_at: depot for picking up empty container. For Maersk bookings: use the location name of the "Empty Container Depot" row from the Load Itinerary table (Page 2).
- return_place: Laden Return / Return to location. For Maersk bookings: use the location name of the "Return Equip Delivery Terminal" row from the Load Itinerary table (Page 2). For LCL shipments, use the value from "Stuffing at" or "Loading at" field instead.
- paperless_code: exact 4-digit number. For Maersk bookings: first find the "Return Equip Delivery Terminal" location from the Load Itinerary table (Page 2), then look up the matching 4-digit code from the "Paperless Code" line on Page 1 — e.g. if return terminal is "Lat Krabang" find the code next to "LKB" (e.g. "B1/LKB/TICT 2811" → 2811); if "Sahathai" or "SHCT" find the code next to "SHCT" (e.g. "SHCT 0520" → 0520); if "TICT" find the code next to "TICT". For other carriers: exact 4-digit number from labels "PAPERLESS CODE", "PORT CODE", or inside parentheses like "(KERRY : 2816)" — extract only the number. If the PAPERLESS CODE section lists multiple codes by terminal (e.g. Yang Ming format with lines like "JTC : EX. LCB KERRY TERMINAL = 2816"), look at the "Turn-In At" field to identify which terminal this booking uses, then find the matching 4-digit code from that terminal's entry in the PAPERLESS CODE section.
- container_type: combine ALL container counts and types e.g. "1X40HC+1X20GP" or "2X40HC". For Maersk bookings: convert the Equipment table format — "40 DRY 9 6" = 40HC, "40 DRY" = 40GP, "20 DRY" = 20GP. null if LCL or no container.
- Dates: dd/mm/yyyy. Cut-offs include hh:mm.
- cy_date: Date the empty container is available for pick-up. For Maersk bookings: look at the Load Itinerary table (Page 2) → find the "Empty Container Depot" row → read its "Release Date" column. Convert YYYY-MM-DD to dd/mm/yyyy. For other carriers: use the empty pick-up date field.
- return_date_1st: 1st Return Date / Turn-In Date / Gate-In Date. For Maersk bookings: calculate ETD minus 5 days and use that date. For MSC bookings: use the "First Receiving" date from the DRY row in the GATE-IN AT TERMINAL/DEPOT table.
- liner_cutoff: Gate Closing / Closing Date / CY Cut-off / Last Load. For Maersk bookings: look at the Load Itinerary table (Page 2) → find the "Return Equip Delivery Terminal" row → read the location name → match EXACTLY to the cut-off on Page 1 using this mapping: "Lat Krabang" → "Cut-Off (DRY and REEF) Lat Krabang"; "TICT" → "Cut-Off TICT"; "Sahathai"/"SHCT" → "Cut-Off SHCT (Sahathai)"; "Laem Chabang" → "Cut-Off (DRY) Laem Chabang". The location name must match exactly — "Lat Krabang" is NOT the same as "TICT". Example: Return terminal = "Lat Krabang" → correct answer is "Cut-Off (DRY and REEF) Lat Krabang" = 05/04/2026 22:00, NOT Cut-Off TICT 06/04/2026 10:00. Ignore DG and REEF-only cutoffs.
- si_cutoff: SI Cut-off / Doc Cut-off / Shipping Particular Cut-off. For Maersk bookings: use "SI (Transshipment & Intra-Asia)" or "SI (Direct)" line from Page 1.
- vgm_cutoff: VGM line.
- If a cut-off shows only a weekday (e.g. "THU"), calculate the actual date from the document date or ETD.
- If ETA is given as a range (e.g. "19/May/2026 - 22/May/2026"), use the first date.
- ETD must be earlier than ETA.
- For MSC bookings: the "EST. TIME OF ARRIVAL/DEPARTURE" field shows two dates — use the SECOND date as ETD (the first date is vessel arrival at POL, the second is departure from POL).
- null if not found."""


def extract_from_pdf(file_bytes: bytes) -> list[dict]:
    """ส่ง PDF ให้ Gemini อ่านครั้งเดียว (รวม dates + general)"""
    ai_config = types.GenerateContentConfig(
        response_mime_type="application/json",
        temperature=0.0,
        seed=42,
    )
    res = genai_client.models.generate_content(
        model=GEMINI_MODEL_BOOKING,
        contents=[
            types.Content(
                role="user",
                parts=[
                    types.Part.from_text(text=PROMPT_BOOKING),
                    types.Part.from_bytes(
                        data=file_bytes,
                        mime_type="application/pdf",
                    ),
                ],
            )
        ],
        config=ai_config,
    )
    raw = res.text.strip()
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
    result = json.loads(raw.strip())
    return result if isinstance(result, list) else [result]
 
 
def save_to_supabase(data_list: list[dict]) -> bool:
    """Upsert to bookings (by booking_no) and insert to revisions."""
    try:
        df = pd.DataFrame(data_list).replace({pd.NA: None, float("nan"): None})
        df = df.where(pd.notnull(df), None)
        df_clean = df.dropna(subset=["booking_no"]).drop_duplicates(
            subset=["booking_no"], keep="last"
        )
        if df_clean.empty:
            st.warning("⚠️ ไม่พบ Booking No. ในเอกสาร — ไม่ได้บันทึกลงฐานข้อมูล")
            return False
        supabase.table(TBL_BOOKINGS).upsert(df_clean.to_dict(orient="records")).execute()
        supabase.table(TBL_REVISIONS).insert(df.to_dict(orient="records")).execute()
        return True
    except Exception as e:
        st.error(f"❌ Database Error: {e}")
        return False
 
 
def bkk_time(df: pd.DataFrame, col: str) -> pd.DataFrame:
    """Convert a UTC timestamp column to Bangkok time string."""
    if col in df.columns:
        try:
            df[col] = (
                pd.to_datetime(df[col])
                .dt.tz_convert("Asia/Bangkok")
                .dt.strftime("%d/%m/%Y %H:%M")
            )
        except Exception:
            pass
    return df
 
 
def to_excel(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter", engine_kwargs={"options": {"nan_inf_to_errors": True}}) as writer:
        df.to_excel(writer, index=False, sheet_name="Bookings")
        wb  = writer.book
        ws  = writer.sheets["Bookings"]
 
        # Formats
        hdr_fmt = wb.add_format({
            "bold": True, "font_name": "Arial",
            "bg_color": "#FFCC00", "font_color": "#D40511",
            "border": 1, "align": "center", "valign": "vcenter",
        })
        cell_fmt = wb.add_format({
            "font_name": "Arial", "font_size": 10,
            "border": 1, "valign": "vcenter",
        })
        alt_fmt = wb.add_format({
            "font_name": "Arial", "font_size": 10,
            "bg_color": "#FFF9E6", "border": 1, "valign": "vcenter",
        })
 
        # Header row
        for col_idx, col_name in enumerate(df.columns):
            ws.write(0, col_idx, col_name, hdr_fmt)
            ws.set_column(col_idx, col_idx, max(len(str(col_name)) + 4, 14))
 
        # Data rows
        for row_idx in range(len(df)):
            fmt = alt_fmt if row_idx % 2 else cell_fmt
            for col_idx in range(len(df.columns)):
                ws.write(row_idx + 1, col_idx, df.iloc[row_idx, col_idx], fmt)
 
        ws.set_row(0, 22)
        ws.freeze_panes(1, 0)
 
    return buf.getvalue()
 
 
def render_table(df: pd.DataFrame, table_id: str = "main") -> None:
    """แสดงตารางแบบ HTML มี badge สี ควบคุมได้เต็มที่"""
 
    # CSS สำหรับตาราง (ใส่ครั้งแรกครั้งเดียว)
    st.markdown("""
    <style>
    .dhl-table-wrap {
        overflow-x: auto;
        overflow-y: auto;
        max-height: 500px;
        border: 1px solid #e8e8e8;
        border-radius: 12px;
        margin-bottom: 8px;
        box-shadow: 0 1px 4px rgba(0,0,0,0.05);
    }
    .dhl-table {
        width: 100%;
        border-collapse: collapse;
        font-size: 12px;
        font-family: 'Inter', sans-serif;
        background: #ffffff;
    }
    .dhl-table thead tr {
        background: #fafafa;
        border-bottom: 1px solid #e8e8e8;
        position: sticky;
        top: 0;
        z-index: 1;
    }
    .dhl-table thead th {
        padding: 11px 14px;
        text-align: left;
        font-size: 10px;
        font-weight: 600;
        color: #aaa;
        letter-spacing: 1px;
        text-transform: uppercase;
        white-space: nowrap;
    }
    .dhl-table tbody tr {
        border-bottom: 1px solid #f0f0f0;
        transition: background 0.1s;
    }
    .dhl-table tbody tr:hover { background: #fdf5f5; }
    .dhl-table tbody td {
        padding: 9px 14px;
        white-space: nowrap;
        color: #444;
    }
    .dhl-num    { color: #ccc !important; }
    .dhl-bkno   { color: #D40511 !important; font-weight: 600; }
    .dhl-code   { color: #D40511 !important; font-weight: 700; letter-spacing: 1px; }
    .dhl-none   { color: #ddd !important; font-style: italic; }
    .dhl-badge  {
        display: inline-block;
        padding: 2px 9px;
        border-radius: 99px;
        font-size: 10px;
        font-weight: 600;
    }
    .b-fcl   { background: #eff6ff; color: #2563eb; }
    .b-lcl   { background: #faf5ff; color: #7c3aed; }
    .b-icd   { background: #eff6ff; color: #3b82f6; }
    .b-alpha { background: #f0fdf4; color: #16a34a; }
    .b-type  { background: #fffbeb; color: #d97706; border: 1px solid #fde68a; }
    .b-sea   { background: #f0f9ff; color: #0284c7; }
    .b-air   { background: #fff7ed; color: #ea580c; }
    .dhl-table-footer {
        padding: 9px 14px;
        border-top: 1px solid #f0f0f0;
        font-size: 11px;
        color: #bbb;
        background: #fafafa;
        border-radius: 0 0 12px 12px;
    }
    </style>
    """, unsafe_allow_html=True)
 
    def val(row, key):
        v = row.get(key, None)
        if v is None or str(v).strip() in ("", "None", "nan", "NaN"):
            return None
        if isinstance(v, float) and v == int(v):
            return str(int(v))
        return str(v).strip()
 
    def badge(text, cls):
        return f'<span class="dhl-badge {cls}">{text}</span>'
 
    rows_html = ""
    for i, (_, row) in enumerate(df.iterrows(), start=1):
        row = row.to_dict()
 
        # booking_no
        bkno  = val(row, "booking_no")
        bkno_html = f'<td class="dhl-bkno">{bkno}</td>' if bkno else '<td class="dhl-none">—</td>'
 
        # loading_at badge
        wh    = val(row, "loading_at") or ""
        wh_cls = "b-icd" if "ICD" in wh else "b-alpha"
        wh_html = badge(wh, wh_cls) if wh else "—"
 
        # FCL/LCL badge
        fcl   = val(row, "fcl_or_lcl") or ""
        fcl_cls = "b-fcl" if fcl == "FCL" else "b-lcl"
        fcl_html = badge(fcl, fcl_cls) if fcl else "—"
 
        # Sea/Air badge
        mode  = val(row, "by_air_or_sea") or ""
        mode_cls = "b-sea" if mode == "Sea" else "b-air"
        mode_html = badge(mode, mode_cls) if mode else "—"
 
        # container_type badge
        ctype = val(row, "container_type") or ""
        ctype_html = badge(ctype, "b-type") if ctype else '<span class="dhl-none">—</span>'
 
        # paperless_code
        code  = val(row, "paperless_code")
        code_html = f'<span class="dhl-code">{code}</span>' if code else '<span class="dhl-none">—</span>'
 
        def cell(key):
            v = val(row, key)
            return f'<td>{v}</td>' if v else '<td class="dhl-none">—</td>'
 
        rows_html += f"""
        <tr>
            <td class="dhl-num">{i}</td>
            {bkno_html}
            <td>{wh_html}</td>
            <td>{fcl_html}</td>
            <td>{mode_html}</td>
            {cell("country")}
            {cell("port_of_destination")}
            {cell("liner_name")}
            {cell("vessel_name")}
            {cell("no_container")}
            <td>{ctype_html}</td>
            {cell("no_pallet")}
            {cell("etd")}
            {cell("eta")}
            {cell("liner_cutoff")}
            {cell("vgm_cutoff")}
            {cell("si_cutoff")}
            {cell("cy_date")}
            {cell("cy_at")}
            {cell("return_date_1st")}
            {cell("return_place")}
            <td>{code_html}</td>
            {cell("updated_at")}
        </tr>"""
 
    headers = [
        "#", "Booking No.", "Loading At", "FCL/LCL", "Mode",
        "Country", "Port of Dest.", "Liner", "Vessel",
        "Ctrs", "Type", "Pallets",
        "ETD", "ETA", "Liner Cutoff", "VGM Cutoff", "SI Cutoff",
        "CY Date", "CY At", "1st Return", "Return Place",
        "Code", "Updated",
    ]
    thead = "".join(f"<th>{h}</th>" for h in headers)
 
    st.markdown(f"""
    <div class="dhl-table-wrap">
        <table class="dhl-table" id="dhl-{table_id}">
            <thead><tr>{thead}</tr></thead>
            <tbody>{rows_html}</tbody>
        </table>
        <div class="dhl-table-footer">แสดง {len(df)} รายการ</div>
    </div>
    """, unsafe_allow_html=True)
 
 
# ─────────────────────────────────────────
# 5b. LOCAL CHARGES — PROMPT & FUNCTIONS
# ─────────────────────────────────────────
PROMPT_LOCAL_CHARGES = """You are a DHL Logistics Analyst. Extract local charge invoice data (in Thai Baht) from this PDF.

Return ONLY a JSON object (no markdown, no explanation):
{
  "agent_invoice_no": <string or null>,
  "pay_to":       <string or null>,
  "tax_name":     <string or null>,
  "tax_id":       <string or null>,
  "delivery_port": <string or null>,
  "etd":          <string DD/MM/YYYY or null>,
  "bl_no":        <string or null>,
  "due_date":     <string DD/MM/YYYY or null>,
  "vat_applicable": <true or false>,
  "items": [
    {
      "description": <string>,
      "category":    <string>,
      "wht_pct":     <0, 1, or 3>,
      "rate":        <number or null>,
      "qty":         <number or null>,
      "total":       <number>
    }
  ]
}

Rules:
- agent_invoice_no: invoice number issued by the freight forwarder/agent — look for:
  1. Labels such as "Invoice No.", "Invoice Number", "INV No.", "Invoice #" near the top of the document
  2. If no label found, look for a prominent alphanumeric code in the document title or heading (e.g. "INVOICE BKK003521Z" → extract "BKK003521Z")
- pay_to: name of the freight forwarder or agent who issued this invoice — look for the company logo, letterhead, or "From" company at the top-right or bottom of the document.
- tax_name: full name AND full address of the freight forwarder who issued this invoice, combined into one string. The issuer is the company whose logo/letterhead appears on the document — look for their address in the footer or "Service provider" section. For invoices with a Thai agent (e.g. "C/O" or "as agent for"), use the Thai local entity's name and address. NOT the "Invoice To" / "Customer" section at the top.
- tax_id: tax identification number of the issuing freight forwarder — search in the footer, bottom of page, or near the issuer's company name/address. It may appear as "Tax ID", "TAX ID", "เลขประจำตัวผู้เสียภาษี", or an unlabeled number near the issuer's details. Do NOT use the tax ID from "Invoice To" / "Customer" / "Billed To" section at the top (that is the recipient's tax ID).
- delivery_port: port of delivery or destination port — format as "Port Name, Country" e.g. "Mombasa, Kenya"
- etd: estimated time of departure in DD/MM/YYYY format
- due_date: payment due date — look for labels "Due Date", "Payment Due", "Due", "วันครบกำหนดชำระ". Format as DD/MM/YYYY. null if not found.
- bl_no: Bill of Lading number — extract using this priority:
  1. Kuehne+Nagel: use "KN TRACKING NO."
  2. Others: "House Bill of Lading" or "House B/L" first
  3. Fallback: "B/L No.", "Bill of Lading", or "OBL NO."
  4. Never use "Master Bill of Lading" or "MB/L"
- vat_applicable: true if invoice mentions "7% VAT", "VAT 7%", "7.00% PURSUANT TO SECTION 80 (2) OF TRC" or has a VAT line item. false otherwise.
- items: list of ALL charge line items found in the invoice (exclude VAT and WHT rows — those are calculated by the system).
  - description: exact charge name as shown in invoice
  - category: classify this charge into one of these fixed values:
      "thc_40hc"        → Terminal Handling Charge for 40HC container
      "thc_40dv"        → Terminal Handling Charge for 40DV/40GP container
      "thc_20gp"        → Terminal Handling Charge for 20GP container
      "export_handling" → Export Handling / Handling Origin
      "seal"            → Seal Fee
      "bl_fee"          → B/L Fee / Bill of Lading Fee
      "surrender_fee"   → Surrender Fee / Telex Release
      "vgm_fee"         → VGM Fee / VGM Submission / VGM Coordination
      "doc_amendment"   → Documentation Amendment / Doc Amendment
      "detention"       → Detention
      "demurrage"       → Demurrage
      "container_repair"→ Container Repair
      "edi_fee"         → EDI Fee / EDI Transmission
      "late_gate"       → Late Gate / Late Gate Service
      "environmental_fee" → Environmental Fee / Green Fee
      "storage"           → Storage / Storage Fee / Container Storage
      "freight_charge"    → Freight / Freight Charge / Ocean Freight / Air Freight / Sea Freight
      "other"           → anything that does not match the above
  - wht_pct: WHT rate for this item. Determine using this priority:
    1. If the charge has a "(W/H 1%)" or "(W/H 3%)" label next to it → use that rate.
    2. If there is a "WHT IN THB" column with entries like "1%=93.8" or "3%=30" next to the charge:
       - Read the digit BEFORE the % sign as the wht_pct (e.g. "1%=93.8" → wht_pct=1, "3%=30" → wht_pct=3)
       - The number AFTER "=" is the pre-calculated WHT amount — do NOT use it as the item total
       - Item total must come from the CHARGES IN THB column
       - IMPORTANT: If this column exists in the invoice, apply it to ALL items and do NOT use rule 5 (Expeditors) at all — even for SEAL, VGM, HANDLING
    2b. If the table has a "W/T%" column with per-row values like "01" or "03" (e.g. Maritime Alliance format):
       - Read the value for each row as wht_pct ("01" → 1, "03" → 3)
       - This takes priority over rules 3–6. Do NOT use global remarks or issuer-based rules.
    3. If there is no per-item label but the document has a global remark applying WHT to all charges (e.g. "Please deduct 3% withholding tax from total Service Charge", "หัก ณ ที่จ่าย 3%") → apply that rate to ALL items.
    4. If the invoice contains "PLEASE PAY WITHOUT DEDUCTION" or "NO DEDUCTION" → wht_pct = 0 for ALL items. Do NOT apply rule 5 (Expeditors).
    5. If the issuer is Expeditors and no WHT is stated in the invoice, apply based on description keywords:
       - WHT 1%: description contains "THC" (any container type), "B/L", "BL FEE", "BILL OF LADING", or "SURRENDER"
       - WHT 3%: description contains "SEAL", "HANDLING", or "VGM"
       - WHT 0%: all other charges
    5b. If the issuer is CEVA and no per-item WHT is stated → wht_pct = 3 for ALL items.
    5c. If the issuer is DSV and no per-item WHT is stated, apply based on shipment_type context provided:
       - Ocean Export:
         WHT 3%: description is specifically "Export Handling", "Handling Fee", or "Handling Charge" (standalone handling service — NOT Terminal Handling Charge / THC)
         WHT 1%: ALL other charges including THC (Terminal Handling Charge), B/L Fee, Surrender Fee, SEAL, VGM, Environmental Fee, Security Fee, EDI Fee, Late Gate, Detention, Demurrage, and any other non-handling charge
       - Air Export / Air Import / Ocean Import:
         WHT 0%: description contains "FREIGHT" or "OCEAN FREIGHT" or "AIR FREIGHT"
         WHT 1%: description contains "TRANSPORT" or "TRUCKING" or "DELIVERY"
         WHT 3%: ALL other charges (Export Handling, Handling Fee, SEAL, VGM, ENVIRONMENTAL, SECURITY, EDI, etc.)
    6. Default: 0
  - rate: unit rate in THB. If the invoice has a RATE column in foreign currency with an EXCH RATE column, convert: rate = RATE × EXCH RATE. If no rate is shown (flat fee), set rate = total.
  - qty: number of units. If the invoice has an explicit QTY column, always use that value — even if the rate is in a foreign currency. If no quantity is shown (flat fee), set qty = 1.
  - total: total amount for this line item (required)
- Use numeric values only (no currency symbols, no commas). null if not found.
"""


def extract_local_charges(file_bytes: bytes, shipment_type: str = "") -> dict:
    ai_config = types.GenerateContentConfig(
        response_mime_type="application/json",
        temperature=0.0,
        seed=42,
    )
    prompt = PROMPT_LOCAL_CHARGES
    if shipment_type:
        prompt += f"\n\nContext: shipment_type = \"{shipment_type}\" (use this for DSV WHT rule 5b)"
    res = genai_client.models.generate_content(
        model=GEMINI_MODEL,
        contents=[
            types.Content(
                role="user",
                parts=[
                    types.Part.from_text(text=prompt),
                    types.Part.from_bytes(data=file_bytes, mime_type="application/pdf"),
                ],
            )
        ],
        config=ai_config,
    )
    raw = res.text.strip()
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
    result = json.loads(raw.strip())

    # Handle case where AI returns a list (multiple invoices in one PDF)
    if isinstance(result, list):
        result = result[0] if result else {}
        result["_multi_invoice"] = True

    items = result.get("items") or []

    # Calculate VAT 7% in Python
    if result.get("vat_applicable"):
        subtotal = sum(float(it.get("total") or 0) for it in items)
        if subtotal > 0:
            result["vat_7"] = round(subtotal * 0.07, 2)

    # Calculate WHT 1% and 3% from per-item wht_pct
    wht1_sum = sum(float(it.get("total") or 0) for it in items if int(it.get("wht_pct") or 0) == 1)
    wht3_sum = sum(float(it.get("total") or 0) for it in items if int(it.get("wht_pct") or 0) == 3)
    result["wht_1"] = round(wht1_sum * 0.01, 2) if wht1_sum > 0 else None
    result["wht_3"] = round(wht3_sum * 0.03, 2) if wht3_sum > 0 else None

    return result


def save_local_charge_v2(header: dict, items: list, pdf_bytes: bytes = None, filename: str = None) -> bool:
    try:
        # อัพโหลด invoice PDF ไปยัง Supabase Storage ก่อน
        if pdf_bytes and filename:
            import uuid as _uuid
            path = f"{_uuid.uuid4()}_{filename}"
            supabase.storage.from_("local-charge-invoices").upload(
                path, pdf_bytes, {"content-type": "application/pdf"}
            )
            header["invoice_pdf_path"] = path
        res = supabase.table(TBL_LOCAL_CHARGES_V2).insert(header).execute()
        lc_id = res.data[0]["id"]
        for item in items:
            item["local_charge_id"] = lc_id
        supabase.table(TBL_LOCAL_CHARGE_ITEMS).insert(items).execute()
        return True
    except Exception as e:
        st.error(f"❌ Supabase Error: {e}")
        return False


# ─────────────────────────────────────────
# 5c. EXPORT SUMMARY — PDF GENERATOR
# ─────────────────────────────────────────
def generate_expense_pdf(records: list[dict], prepared_by: str = "", prepared_by_phone: str = "") -> bytes:
    """Generate expense summary PDF. records = list of {header, items}"""
    from fpdf import FPDF

    from pathlib import Path
    import tempfile, urllib.request
    _win = Path("C:/Windows/Fonts")
    if _win.exists() and (_win / "tahoma.ttf").exists():
        FONT_PATH    = str(_win / "tahoma.ttf")
        FONT_PATH_BD = str(_win / "tahomabd.ttf")
    else:
        _tmp = Path(tempfile.gettempdir())
        FONT_PATH    = str(_tmp / "Sarabun-Regular.ttf")
        FONT_PATH_BD = str(_tmp / "Sarabun-Bold.ttf")
        if not Path(FONT_PATH).exists():
            urllib.request.urlretrieve("https://github.com/google/fonts/raw/main/ofl/sarabun/Sarabun-Regular.ttf", FONT_PATH)
        if not Path(FONT_PATH_BD).exists():
            urllib.request.urlretrieve("https://github.com/google/fonts/raw/main/ofl/sarabun/Sarabun-Bold.ttf", FONT_PATH_BD)
    LOGO_PATH    = str(Path(__file__).parent / "Logo.png")

    class PDF(FPDF):
        def header(self):
            # Logo top-left — h=14 วางที่ y=10
            self.image(LOGO_PATH, x=10, y=10, h=14)
            # Company info เริ่มที่ x=55 ให้พ้น logo
            self.set_xy(55, 10)
            self.set_font("Tahoma", "B", 9)
            self.cell(0, 5, "DHL SUPPLY CHAIN (THAILAND)", new_x="LMARGIN", new_y="NEXT")
            self.set_x(55)
            self.set_font("Tahoma", "", 7.5)
            self.cell(0, 4.5, "NO. 9 G TOWER GRAND RAMA 9 (NORTH WING), 26TH FLOOR AND 27TH FLOOR,", new_x="LMARGIN", new_y="NEXT")
            self.set_x(55)
            self.cell(0, 4.5, "RAMA IX RD, HUAI KHWANG, BANGKOK 10310  TEL. (02) 779 9800", new_x="LMARGIN", new_y="NEXT")
            self.ln(5)
            if self.page_no() > 1:
                self.set_font("Tahoma", "B", 13)
                self.cell(0, 8, "EXPENSE DETAIL", new_x="LMARGIN", new_y="NEXT", align="C")
                self.ln(2)

    pdf = PDF()
    pdf.add_font("Tahoma",  "", FONT_PATH)
    pdf.add_font("Tahoma",  "B", FONT_PATH_BD)
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.set_margins(10, 15, 10)

    # column widths: 28+72+25+7+25+23 = 180mm
    CW = {"inv": 28, "name": 72, "rate": 25, "x": 7, "qty": 25, "amt": 23}
    COL_W  = 35
    LABEL_W = CW["inv"] + CW["name"] + CW["rate"] + CW["x"] + CW["qty"]

    def info_row(label, value, multiline=False):
        pdf.set_font("Tahoma", "B", 9)
        pdf.cell(COL_W, 6, label)
        pdf.set_font("Tahoma", "", 9)
        if multiline:
            pdf.multi_cell(0, 6, str(value or ""), new_x="LMARGIN", new_y="NEXT")
        else:
            pdf.cell(0, 6, str(value or ""), new_x="LMARGIN", new_y="NEXT")

    # Summary row colors matching sample PDF
    COLOR = {
        "amount": (255, 213, 128),   # amber
        "vat":    (226, 239, 218),   # light green
        "total":  (226, 239, 218),   # light green
        "wht":    (255, 255, 153),   # light yellow
        "net":    (189, 215, 238),   # light blue
    }

    def sum_row(label, value, color_key="amount"):
        r, g, b = COLOR[color_key]
        bold = color_key == "net"
        pdf.set_font("Tahoma", "B" if bold else "", 8)
        pdf.set_fill_color(r, g, b)
        pdf.cell(LABEL_W, 6, label, border=1, fill=True, align="R")
        pdf.cell(CW["amt"], 6, f"{value:,.2f}", border=1, fill=True, align="R", new_x="LMARGIN", new_y="NEXT")

    # ── Cover Page ──────────────────────────────────────────
    # total width = 40+28+24+32+20+20+21 = 185mm (fits A4 190mm printable)
    CW_COV = {"part": 40, "inv": 28, "country": 24, "payto": 32, "due": 20, "remark": 20, "amt": 21}
    COV_LINE_H = 5  # height per line in cover table

    def _count_lines(pdf_obj, text, col_w):
        """Count lines needed for text in a given column width."""
        if not text:
            return 1
        words = str(text).split()
        lines, line_w = 1, 0.0
        for word in words:
            ww = pdf_obj.get_string_width(word + " ")
            if line_w + ww > col_w - 2 and line_w > 0:
                lines += 1
                line_w = ww
            else:
                line_w += ww
        return lines

    def _draw_cover_row(pdf_obj, cw, vals, aligns, fill=False, bold=False, fill_color=None):
        """Draw one row with auto row-height and word-wrap for part/payto/remark."""
        WRAP_KEYS = {"part", "payto", "remark"}
        if fill_color:
            pdf_obj.set_fill_color(*fill_color)
        pdf_obj.set_font("Tahoma", "B" if bold else "", 7.5)

        # Calculate row height
        max_lines = 1
        for key in WRAP_KEYS:
            if key in cw:
                max_lines = max(max_lines, _count_lines(pdf_obj, vals.get(key, ""), cw[key]))
        row_h = max_lines * COV_LINE_H

        x0, y0 = pdf_obj.get_x(), pdf_obj.get_y()
        x = x0
        for key, w in cw.items():
            text = str(vals.get(key) or "")
            align = aligns.get(key, "L")
            pdf_obj.set_xy(x, y0)
            if key in WRAP_KEYS:
                # วาด background fill ก่อน (ถ้ามี)
                if fill and fill_color:
                    pdf_obj.set_fill_color(*fill_color)
                    pdf_obj.rect(x, y0, w, row_h, style="F")
                # วาด border รอบ cell ด้วย rect
                pdf_obj.rect(x, y0, w, row_h)
                # วาด text ด้วย multi_cell โดยไม่มี border (center แนวตั้งเหมือน cell())
                n_lines = _count_lines(pdf_obj, text, w)
                v_offset = max(0, (row_h - n_lines * COV_LINE_H) / 2)
                pdf_obj.set_xy(x + 1, y0 + v_offset)
                pdf_obj.multi_cell(w - 2, COV_LINE_H, text, border=0, align=align,
                                   fill=False, new_x="RIGHT", new_y="TOP")
            else:
                pdf_obj.cell(w, row_h, text, border=1, align=align, fill=fill)
            x += w
        pdf_obj.set_xy(x0, y0 + row_h)

    COV_ALIGNS = {"part": "L", "inv": "C", "country": "C", "payto": "L", "due": "C", "remark": "L", "amt": "R"}

    pdf.add_page()
    pdf.set_font("Tahoma", "B", 14)
    pdf.cell(0, 8, "ใบแจ้งค่าใช้จ่ายส่งออก", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.set_font("Tahoma", "B", 12)
    pdf.cell(0, 7, "EXPORT EXPENSE DETAIL", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(6)

    today_cover = datetime.now(pytz.timezone("Asia/Bangkok")).strftime("%d/%m/%Y")
    pdf.set_font("Tahoma", "", 9)
    pdf.cell(0, 6, f"Date : {today_cover}", new_x="LMARGIN", new_y="NEXT", align="R")
    pdf.ln(4)

    # Cover table header
    pdf.set_fill_color(220, 220, 220)
    pdf.set_font("Tahoma", "B", 7.5)
    hdr_vals = {"part": "รายการ / PARTICULARS", "inv": "INVOICE NO.", "country": "Country",
                "payto": "Pay To", "due": "Due Date", "remark": "Remark", "amt": "Amount"}
    _draw_cover_row(pdf, CW_COV, hdr_vals, {k: "C" for k in CW_COV}, fill=True, bold=True, fill_color=(220, 220, 220))

    cover_total = 0.0
    for rec in records:
        hdr  = rec["header"]
        bk   = rec.get("bk") or {}
        its  = rec["items"]
        subtotal = sum(float(it.get("total") or 0) for it in its)
        vat_7 = float(hdr.get("vat_7") or 0)
        wht_1 = float(hdr.get("wht_1") or 0)
        wht_3 = float(hdr.get("wht_3") or 0)
        net   = round(subtotal + vat_7 - wht_1 - wht_3, 2)
        cover_total += net

        cats = sorted(set(it.get("category") or "other" for it in its))
        default_part = "+ ".join(c.upper().replace("_", "") for c in cats)
        cat_label = rec.get("cover_part") or default_part

        row_vals = {
            "part":    cat_label,
            "inv":     rec.get("cover_inv")     or hdr.get("ctc_invoice_no") or "-",
            "country": rec.get("cover_country") or bk.get("country") or "-",
            "payto":   rec.get("cover_payto")   or hdr.get("pay_to") or "-",
            "due":     rec.get("cover_due")     or hdr.get("due_date") or "-",
            "remark":  rec.get("cover_remark")  or hdr.get("remark") or "",
            "amt":     f"{net:,.2f}",
        }
        _draw_cover_row(pdf, CW_COV, row_vals, COV_ALIGNS)

    # Total row
    total_label_w = sum(CW_COV[k] for k in ["part", "inv", "country", "payto", "due", "remark"])
    pdf.set_font("Tahoma", "B", 8)
    pdf.set_fill_color(189, 215, 238)
    pdf.cell(total_label_w, COV_LINE_H, "Total", border=1, fill=True, align="R")
    pdf.cell(CW_COV["amt"], COV_LINE_H, f"{cover_total:,.2f}", border=1, fill=True, align="R", new_x="LMARGIN", new_y="NEXT")

    pdf.ln(6)
    pdf.set_font("Tahoma", "", 9)
    pdf.ln(10)
    name_line = f"Prepared by : {prepared_by}" if prepared_by else "Prepared by ............................................"
    pdf.cell(80, 6, name_line)
    pdf.cell(80, 6, "Approved by ............................................", new_x="LMARGIN", new_y="NEXT")
    if prepared_by_phone:
        pdf.ln(2)
        pdf.set_font("Tahoma", "", 8)
        pdf.cell(80, 5, f"Tel. : {prepared_by_phone}")

    # ── Detail pages ────────────────────────────────────────
    for rec in records:
        hdr   = rec["header"]
        items = rec["items"]

        pdf.add_page()

        # ── Header info ──
        today_str = datetime.now(pytz.timezone("Asia/Bangkok")).strftime("%d/%m/%Y")

        # DATE: วางที่ขวาบนแบบ absolute (y=10 ระดับเดียวกับ logo/company info)
        y_after_header = pdf.get_y()
        pdf.set_xy(142, 10)
        pdf.set_font("Tahoma", "B", 9)
        pdf.cell(28, 5, "DATE :", align="R")
        pdf.set_font("Tahoma", "", 9)
        pdf.cell(30, 5, today_str, align="R")
        pdf.set_xy(10, y_after_header)

        # SHIPPER row (ไม่มี DATE แล้ว)
        pdf.set_font("Tahoma", "B", 9)
        pdf.cell(COL_W, 6, "SHIPPER :")
        pdf.set_font("Tahoma", "", 9)
        pdf.cell(0, 6, "CARRIER AIR CONDITIONING (THAILAND) CO.,LTD.", new_x="LMARGIN", new_y="NEXT")
        info_row("PAY TO :", hdr.get("pay_to") or "")
        info_row("TAX NAME :", hdr.get("tax_name") or "", multiline=True)
        pdf.ln(2)
        info_row("TAX ID NO. :", hdr.get("tax_id") or "")
        info_row("DELIVERY PORT :", hdr.get("delivery_port") or "")
        info_row("ETD :", hdr.get("etd") or "")
        info_row("BL NO :", hdr.get("bl_no") or "")
        info_row("BOOKING NO. :", hdr.get("booking_no") or "")
        pdf.ln(3)

        # ── Table header ──
        pdf.set_fill_color(220, 220, 220)
        pdf.set_font("Tahoma", "B", 8)
        pdf.cell(CW["inv"],  7, "INVOICE NO.", border=1, fill=True, align="C")
        pdf.cell(CW["name"], 7, "DESCRIPTION", border=1, fill=True, align="C")
        pdf.cell(CW["rate"], 7, "RATE",        border=1, fill=True, align="C")
        pdf.cell(CW["x"],    7, "x",           border=1, fill=True, align="C")
        pdf.cell(CW["qty"],  7, "QUANTITY",    border=1, fill=True, align="C")
        pdf.cell(CW["amt"],  7, "AMOUNT",      border=1, fill=True, align="C", new_x="LMARGIN", new_y="NEXT")

        # ── Table rows ──
        invoice_no = hdr.get("ctc_invoice_no") or ""
        subtotal = 0.0
        pdf.set_font("Tahoma", "", 8)
        _CAT_ORDER = [
            "thc_40hc","thc_40dv","thc_20gp","export_handling","seal","bl_fee",
            "surrender_fee","vgm_fee","doc_amendment","detention","demurrage",
            "container_repair","edi_fee","late_gate","environmental_fee",
            "storage","freight_charge","other",
        ]
        items = sorted(items, key=lambda x: _CAT_ORDER.index(x.get("category") or "other")
                       if (x.get("category") or "other") in _CAT_ORDER else len(_CAT_ORDER))
        for it in items:
            desc      = it.get("description") or ""
            rate      = float(it.get("rate") or 0)
            qty       = float(it.get("qty") or 0)
            total     = float(it.get("total") or 0)
            subtotal += total

            rate_str  = f"{rate:,.2f}" if rate else ""
            qty_str   = f"{qty:,.3f}" if qty else ""
            total_str = f"{total:,.2f}" if total else "-"

            pdf.cell(CW["inv"],  6, invoice_no, border=1, align="C")
            pdf.cell(CW["name"], 6, desc[:45],  border=1)
            pdf.cell(CW["rate"], 6, rate_str,   border=1, align="R")
            pdf.cell(CW["x"],    6, "x",        border=1, align="C")
            pdf.cell(CW["qty"],  6, qty_str,    border=1, align="R")
            pdf.cell(CW["amt"],  6, total_str,  border=1, align="R", new_x="LMARGIN", new_y="NEXT")
            invoice_no = ""  # show only on first row

        # ── Summary rows ──
        vat_7     = float(hdr.get("vat_7")  or 0)
        wht_1     = float(hdr.get("wht_1")  or 0)
        wht_3     = float(hdr.get("wht_3")  or 0)
        total_net = subtotal + vat_7 - wht_1 - wht_3

        sum_row("Amount",              subtotal,          color_key="amount")
        sum_row("บวก vat 7%",          vat_7,             color_key="vat")
        sum_row("Total",               subtotal + vat_7,  color_key="total")
        sum_row("หัก ภาษี ณ ที่จ่าย 1%", wht_1,          color_key="wht")
        sum_row("หัก ภาษี ณ ที่จ่าย 3%", wht_3,          color_key="wht")
        sum_row("ยอดจ่ายจริง",         total_net,         color_key="net")

    summary_bytes = bytes(pdf.output())

    # ── Merge invoice PDFs ก่อน summary ของแต่ละ invoice ──
    try:
        from pypdf import PdfWriter, PdfReader
        import io as _io

        writer = PdfWriter()

        # page 0 = cover page, detail pages เริ่มที่ index 1
        sum_reader = PdfReader(_io.BytesIO(summary_bytes))
        writer.add_page(sum_reader.pages[0])  # cover page

        for i, rec in enumerate(records):
            # detail page อยู่ที่ index i+1 (เพราะ page 0 = cover)
            writer.add_page(sum_reader.pages[i + 1])
            # แทรก invoice PDF ต่อท้าย (ถ้ามี)
            inv_path = rec["header"].get("invoice_pdf_path")
            if inv_path:
                try:
                    inv_bytes = supabase.storage.from_("local-charge-invoices").download(inv_path)
                    inv_reader = PdfReader(_io.BytesIO(inv_bytes))
                    for page in inv_reader.pages:
                        writer.add_page(page)
                except Exception:
                    pass  # ถ้าดาวน์โหลดไม่ได้ ข้ามไป

        out = _io.BytesIO()
        writer.write(out)
        return out.getvalue()
    except Exception:
        # fallback: return summary only
        return summary_bytes


# ─────────────────────────────────────────
# 6. SIDEBAR NAVIGATION
# ─────────────────────────────────────────
with st.sidebar:
    st.markdown("### 🚚 DHL Logistics Menu")
    page = st.radio(
        "เลือกเมนู",
        ["📤 Upload & Extract", "📄 Generate SI (Draft)", "💰 Local Charges", "📊 Export Summary"],
        label_visibility="collapsed",
    )
    st.divider()
    st.markdown(
        "<span style='font-size:11px;color:#555;'>Powered by Ship Co. (CTC Site)<br>"
        "© DHL Supply Chain Thailand</span>",
        unsafe_allow_html=True,
    )
 
# ─────────────────────────────────────────
# 7. PAGE: UPLOAD & EXTRACT
# ─────────────────────────────────────────
if page == "📤 Upload & Extract":
 
    # ── Upload zone ──────────────────────
    if "uploader_key" not in st.session_state:
        st.session_state["uploader_key"] = 0
 
    col1, col2 = st.columns(2)

    with col1:
        st.info("🏢 **ICD Warehouse**")
        files_icd = st.file_uploader(
            "โยนไฟล์สำหรับ ICD ที่นี่",
            type="pdf",
            accept_multiple_files=True,
            key=f"icd_{st.session_state['uploader_key']}",
        )

    with col2:
        st.success("🏗️ **ALPHA Warehouse**")
        files_alpha = st.file_uploader(
            "โยนไฟล์สำหรับ ALPHA ที่นี่",
            type="pdf",
            accept_multiple_files=True,
            key=f"alpha_{st.session_state['uploader_key']}",
        )

    # ── Auto-process when files are uploaded ─────────────────
    from concurrent.futures import ThreadPoolExecutor, as_completed

    list_icd   = files_icd   or []
    list_alpha = files_alpha or []
    total      = len(list_icd) + len(list_alpha)

    if total > 0:
        all_data     = []
        progress_bar = st.progress(0)
        status_text  = st.empty()
        processed    = 0

        # อ่าน bytes ทั้งหมดใน main thread ก่อน (UploadedFile ไม่ thread-safe)
        tasks = (
            [(f.name, f.read(), "ICD")   for f in list_icd] +
            [(f.name, f.read(), "ALPHA") for f in list_alpha]
        )

        def process_file(name, file_bytes, warehouse):
            items = extract_from_pdf(file_bytes)
            if items:
                for item in items:
                    item["source_file"] = name
                    item["loading_at"]  = warehouse
            return items or []

        with ThreadPoolExecutor(max_workers=min(total, 5)) as executor:
            futures = {
                executor.submit(process_file, name, fb, wh): name
                for name, fb, wh in tasks
            }
            for future in as_completed(futures):
                name = futures[future]
                try:
                    items = future.result()
                    all_data.extend(items)
                except Exception as e:
                    st.error(f"❌ {name}: {e}")
                processed += 1
                progress_bar.progress(processed / total)
                status_text.caption(f"สแกนแล้ว {processed}/{total} ไฟล์ — {name}")

        status_text.empty()

        if all_data:
            if save_to_supabase(all_data):
                st.success(
                    f"🎉 บันทึกข้อมูล {len(all_data)} รายการ "
                    f"จากทั้งหมด {total} ไฟล์ เรียบร้อยแล้ว"
                )
                st.session_state["uploader_key"] += 1
                st.rerun()
 
    # ── Live View ────────────────────────
    st.divider()
 
    st.subheader("📊 รายการ Booking ทั้งหมด (Live View)")
 
    try:
        res = (
            supabase.table(TBL_BOOKINGS)
            .select("*")
            .order("updated_at", desc=True)
            .execute()
        )
 
        if res.data:
            df_live = pd.DataFrame(res.data)
            df_live = bkk_time(df_live, "updated_at")
            existing = [c for c in COLUMNS_ORDER if c in df_live.columns]
            df_show  = df_live[existing].copy()
 
            # Metrics row
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("📦 Total Bookings", len(df_show))
            m2.metric("🏢 ICD",   int(df_show["loading_at"].str.contains("ICD",   na=False).sum()) if "loading_at" in df_show else 0)
            m3.metric("🏗️ ALPHA", int(df_show["loading_at"].str.contains("ALPHA", na=False).sum()) if "loading_at" in df_show else 0)
            m4.metric("🌊 FCL",   int(df_show["fcl_or_lcl"].str.contains("FCL",   na=False).sum()) if "fcl_or_lcl"  in df_show else 0)
 
            st.markdown("<br>", unsafe_allow_html=True)
 
            search = st.text_input("🔍 ค้นหา...", placeholder="Booking No., Port, Vessel, Country...")
 
            if search:
                mask = df_show.astype(str).apply(
                    lambda x: x.str.contains(search, case=False, na=False)
                ).any(axis=1)
                df_show = df_show[mask]
 
            df_show.index = range(1, len(df_show) + 1)
            render_table(df_show, table_id="live")
 
            col_dl1, col_dl2 = st.columns([1, 5])
            with col_dl1:
                st.download_button(
                    "📥 Export Excel",
                    data=to_excel(df_show),
                    file_name="DHL_Bookings.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

            # ── Edit Booking ──────────────────────────────────────
            st.divider()
            st.subheader("✏️ แก้ไขข้อมูล Booking")

            all_bk_nos = [r.get("booking_no") for r in res.data if r.get("booking_no")]
            edit_bk = st.selectbox("เลือก Booking ที่ต้องการแก้ไข",
                                   ["-- เลือก --"] + all_bk_nos, key="edit_bk_select")

            if edit_bk != "-- เลือก --":
                row_data = next((r for r in res.data if r.get("booking_no") == edit_bk), {})

                with st.form("edit_booking_form"):
                    st.markdown(f"**Booking No.: {edit_bk}**")
                    ec1, ec2, ec3 = st.columns(3)

                    loading_at     = ec1.selectbox("Loading At",    ["ICD", "ALPHA"],
                                                    index=["ICD","ALPHA"].index(row_data.get("loading_at","ICD"))
                                                    if row_data.get("loading_at") in ["ICD","ALPHA"] else 0)
                    fcl_or_lcl     = ec2.selectbox("FCL/LCL",       ["FCL","LCL"],
                                                    index=["FCL","LCL"].index(row_data.get("fcl_or_lcl","FCL"))
                                                    if row_data.get("fcl_or_lcl") in ["FCL","LCL"] else 0)
                    by_air_or_sea  = ec3.selectbox("Mode",          ["Sea","Air"],
                                                    index=["Sea","Air"].index(row_data.get("by_air_or_sea","Sea"))
                                                    if row_data.get("by_air_or_sea") in ["Sea","Air"] else 0)

                    ec4, ec5, ec6 = st.columns(3)
                    country        = ec4.text_input("Country",           value=row_data.get("country") or "")
                    port_of_dest   = ec5.text_input("Port of Dest.",     value=row_data.get("port_of_destination") or "")
                    liner_name     = ec6.text_input("Liner",             value=row_data.get("liner_name") or "")

                    ec7, ec8 = st.columns([2, 1])
                    vessel_name    = ec7.text_input("Vessel / Voyage",   value=row_data.get("vessel_name") or "")
                    paperless_code = ec8.text_input("Paperless Code",    value=row_data.get("paperless_code") or "")

                    ec9, ec10, ec11 = st.columns(3)
                    no_container   = ec9.number_input("No. Container",   min_value=0,
                                                       value=int(row_data.get("no_container") or 0))
                    container_type = ec10.text_input("Container Type",   value=row_data.get("container_type") or "")
                    no_pallet      = ec11.number_input("No. Pallet",     min_value=0,
                                                        value=int(row_data.get("no_pallet") or 0))


                    ec12, ec13 = st.columns(2)
                    cy_at          = ec12.text_input("CY At",            value=row_data.get("cy_at") or "")
                    return_place   = ec13.text_input("Return Place",     value=row_data.get("return_place") or "")

                    st.markdown("**วันที่สำคัญ**")
                    ed1, ed2, ed3 = st.columns(3)
                    etd            = ed1.text_input("ETD (dd/mm/yyyy)",  value=row_data.get("etd") or "")
                    eta            = ed2.text_input("ETA (dd/mm/yyyy)",  value=row_data.get("eta") or "")
                    cy_date        = ed3.text_input("CY Date",           value=row_data.get("cy_date") or "")

                    ed4, ed5, ed6 = st.columns(3)
                    liner_cutoff   = ed4.text_input("Liner Cutoff",      value=row_data.get("liner_cutoff") or "")
                    vgm_cutoff     = ed5.text_input("VGM Cutoff",        value=row_data.get("vgm_cutoff") or "")
                    si_cutoff      = ed6.text_input("SI Cutoff",         value=row_data.get("si_cutoff") or "")

                    ed7, ed8 = st.columns(2)
                    return_date    = ed7.text_input("1st Return Date",   value=row_data.get("return_date_1st") or "")

                    submitted = st.form_submit_button("💾 บันทึกการแก้ไข", use_container_width=True)

                if submitted:
                    update_payload = {
                        "booking_no":        edit_bk,
                        "loading_at":        loading_at,
                        "fcl_or_lcl":        fcl_or_lcl,
                        "by_air_or_sea":     by_air_or_sea,
                        "country":           country or None,
                        "port_of_destination": port_of_dest or None,
                        "liner_name":        liner_name or None,
                        "vessel_name":       vessel_name or None,
                        "paperless_code":    paperless_code or None,
                        "no_container":      no_container or None,
                        "container_type":      container_type or None,
                        "no_pallet":           no_pallet or None,
                        "cy_at":             cy_at or None,
                        "return_place":      return_place or None,
                        "etd":               etd or None,
                        "eta":               eta or None,
                        "cy_date":           cy_date or None,
                        "liner_cutoff":      liner_cutoff or None,
                        "vgm_cutoff":        vgm_cutoff or None,
                        "si_cutoff":         si_cutoff or None,
                        "return_date_1st":   return_date or None,
                    }
                    try:
                        supabase.table(TBL_BOOKINGS).upsert(update_payload).execute()
                        st.success(f"✅ บันทึก {edit_bk} เรียบร้อยแล้ว")
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Save Error: {e}")

        else:
            st.info("📌 ยังไม่มีข้อมูล — อัปโหลด PDF เพื่อเริ่มต้น")
 
    except Exception as e:
        st.error(f"Load Error: {e}")
 
    # ── Revision History ─────────────────
    st.divider()
    st.subheader("📜 ประวัติการบันทึกย้อนหลัง (Revision Logs)")
 
    if st.button("🔍 โหลดประวัติทั้งหมด"):
        st.session_state["show_history"] = True
 
    if st.session_state.get("show_history"):
        try:
            rev = (
                supabase.table(TBL_REVISIONS)
                .select("*")
                .order("created_at", desc=True)
                .execute()
            )
            if rev.data:
                df_rev = pd.DataFrame(rev.data)
                df_rev = bkk_time(df_rev, "created_at")
                hist_cols = [c for c in COLUMNS_ORDER if c in df_rev.columns] + (
                    ["created_at"] if "created_at" in df_rev.columns else []
                )
                df_rev = df_rev[hist_cols]
 
                s_hist = st.text_input("🔍 ค้นหาในประวัติ...", key="hist_search")
                if s_hist:
                    mask = df_rev.astype(str).apply(
                        lambda x: x.str.contains(s_hist, case=False, na=False)
                    ).any(axis=1)
                    df_rev = df_rev[mask]
 
                df_rev.index = range(1, len(df_rev) + 1)
                render_table(df_rev, table_id="history")
                st.download_button(
                    "📥 Export History",
                    data=to_excel(df_rev),
                    file_name="DHL_Bookings_History.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            else:
                st.info("ยังไม่มีประวัติ")
        except Exception as e:
            st.error(f"History Error: {e}")
 
# ─────────────────────────────────────────
# 8. PAGE: GENERATE SI
# ─────────────────────────────────────────
elif page == "📄 Generate SI (Draft)":
 
    # ── imports เพิ่มเติมสำหรับ Auto SI ──────────────────────
    import copy
    from openpyxl import load_workbook
 
    PROMPT_INVOICE = """You are a DHL logistics expert. Extract shipping info from this Invoice & Packing List PDF.
Return ONLY a JSON array — one object per INVOICE found. No markdown, no explanation.
[{
  "invoice_no": "e.g. 1075863",
  "description": "SHORT description of goods",
  "shipping_mark": "ALL lines under CASE MARK / SHIPPING MARK section joined with \\n, e.g. 'PO#5400025474\\nHS CODE : 841590'. Copy every line exactly as printed. Do NOT add HS CODE if it is not written in that section.",
  "cartons": "1,389 CARTONS or 40 PP.PALLETS — full package count with unit from TOTAL row",
  "quantity_str": "1,389 SETS or 98,470 PCS — quantity with unit",
  "net_weight_kgs": 33053.00,
  "gross_weight_kgs": 37221.00,
  "measurement_cbm": 282.895,
  "hs_code": "8415.10 or null",
  "consignee_name": "ACCOUNTEE name",
  "consignee_address": "full address, use \\n for line breaks",
  "ship_to_name": "SHIP TO name",
  "ship_to_address": "full address, use \\n for line breaks",
  "vessel_feeder": "feeder vessel + voyage e.g. X-PRESS ANGLESEY V.26002W",
  "vessel_mother": "mother vessel + voyage e.g. ONE HAMMERSMITH V.088W or null",
  "port_of_loading": "FROM port e.g. LAEM CHABANG, THAILAND",
  "port_of_discharge": "TO port e.g. LE HAVRE, FRANCE",
  "transhipment_port": "VIA port or null",
  "etd": "SAILING ON/OR ABOUT date dd/mm/yyyy",
  "carrier": "CARRIER field e.g. EXPEDITORS/ONE — look for 'CARRIER:' label near vessel/voyage info"
}]
Rules: Extract TOTAL row from Packing List. gross_weight/cbm from PL TOTAL.
For 'cartons': always include the unit word (CARTONS, PP.PALLETS, CTNS, etc.) not just the number.
null if not found."""
 
    def _extract_invoices(file_bytes):
        ai_cfg = types.GenerateContentConfig(
            response_mime_type="application/json", temperature=0.0, seed=42
        )
        res = genai_client.models.generate_content(
            model=GEMINI_MODEL,
            contents=[types.Content(role="user", parts=[
                types.Part.from_text(text=PROMPT_INVOICE),
                types.Part.from_bytes(data=file_bytes, mime_type="application/pdf"),
            ])],
            config=ai_cfg,
        )
        raw = res.text.strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"): raw = raw[4:]
        data = json.loads(raw.strip())
        return data if isinstance(data, list) else [data]
 
    def _fill_si(template_bytes, booking, invoices, containers, extra):
        from openpyxl.cell.cell import MergedCell
        from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
        import copy
 
        wb = load_workbook(io.BytesIO(template_bytes))
        ws = wb["Shipping Particular"]
        s  = lambda v: str(v).strip() if v else ""
        first = invoices[0] if invoices else {}
 
        def safe_write(addr, value):
            cell = ws[addr]
            if not isinstance(cell, MergedCell) and (cell.value is None or str(cell.value).strip() == ""):
                cell.value = value
 
        def safe_write_rc(row, col, value):
            cell = ws.cell(row=row, column=col)
            if not isinstance(cell, MergedCell) and (cell.value is None or str(cell.value).strip() == ""):
                cell.value = value
 
        def copy_row_style(src_row, dst_row):
            """copy style ทุก cell จาก src_row ไป dst_row"""
            for c in range(1, 12):
                src = ws.cell(row=src_row, column=c)
                dst = ws.cell(row=dst_row, column=c)
                if isinstance(src, MergedCell) or isinstance(dst, MergedCell):
                    continue
                dst.font      = copy.copy(src.font)
                dst.border    = copy.copy(src.border)
                dst.alignment = copy.copy(src.alignment)
                dst.fill      = copy.copy(src.fill)
                dst.number_format = src.number_format
            ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
 
        # ── Header ─────────────────────────────────────────────
        safe_write("J7", s(booking.get("booking_no")))
        if extra.get("revised"):
            safe_write("I9", "REVISED")
 
        # ── Consignee rows 15-18 ────────────────────────────────
        con_name   = s(first.get("consignee_name"))
        con_addr   = s(first.get("consignee_address"))
        addr_lines = [l.strip() for l in con_addr.replace("\\n","\n").split("\n") if l.strip()]
        a15_cell = ws["A15"]
        if not isinstance(a15_cell, MergedCell) and (a15_cell.value is None or str(a15_cell.value).strip() == ""):
            safe_write("A15", con_name)
            for i, line in enumerate(addr_lines[:5]):
                safe_write(f"A{16+i}", line)
 
        # ── Notify rows 22-25 ───────────────────────────────────
        notify_name = s(first.get("ship_to_name")) or con_name
        notify_addr = s(first.get("ship_to_address")) or con_addr
        nlines = [l.strip() for l in notify_addr.replace("\\n","\n").split("\n") if l.strip()]
        a22_cell = ws["A22"]
        if not isinstance(a22_cell, MergedCell) and (a22_cell.value is None or str(a22_cell.value).strip() == ""):
            safe_write("A22", notify_name)
            for i, line in enumerate(nlines[:4]):
                safe_write(f"A{23+i}", line)
 
        # ── Vessel / Port (จาก Invoice SAP) ────────────────────
        safe_write("A28", s(first.get("vessel_feeder")))
        safe_write("D28", s(first.get("port_of_loading")) or "LAEM CHABANG, THAILAND")
        safe_write("I28", s(first.get("port_of_loading")) or "LAEM CHABANG, THAILAND")
 
        etd_str = s(first.get("etd"))
        try:
            etd_dt = datetime.strptime(etd_str, "%d/%m/%Y")
            cell_a30 = ws["A30"]
            if not isinstance(cell_a30, MergedCell):
                cell_a30.value = etd_dt
                cell_a30.number_format = "DD-MMM-YY"
        except Exception:
            safe_write("A30", etd_str)
 
        safe_write("D30", s(first.get("port_of_discharge")))
        safe_write("I30", s(first.get("port_of_discharge")))  # Place of Delivery = same
 
        # A32 = Place of Issue (BANGKOK, THAILAND — static)
        safe_write("A32", "BANGKOK, THAILAND")
        # D32 = Transhipment port
        trans = s(first.get("transhipment_port"))
        safe_write("D32", trans + ",SINGAPORE" if trans and "SINGAPORE" in trans.upper() and "," not in trans else trans)
        safe_write("I32", s(first.get("vessel_mother")))
 
        # ── I13 = Carrier (จาก Invoice SAP) ────────────────────
        carrier = s(first.get("carrier"))
        safe_write("I13", carrier)
 
        # ── Container count label ───────────────────────────────
        safe_write("B35", s(booking.get("container_type")) or "")
 
        # ── I36/J36 = KGS. / CBM (หน่วยใต้ตัวเลข total) ───────
        safe_write("I36", "KGS.")
        safe_write("J36", "CBM")
 
        # ── Cargo block ─────────────────────────────────────────
        # Template layout:
        #   row 37: B=CARTONS (total qty label — มีแค่อันเดียว), formula total
        #   row 38: A=mark1, B=CARTONS(หน่วย—มีแค่ invoice แรก), C=qty, D=CARTONS, E=(qty_str)
        #   row 39: A=mark2, C=description
        #   row 40: A=mark3, C=INVOICE NO.
        #   row 41: C=G.W., D=gw, E=KGS, F=M3, G=cbm, H=CBM
        #   row 42: (blank spacer)
        #   row 43: invoice 2 qty row (ไม่มี B=CARTONS แล้ว)
        #   ...
 
        marks = list(dict.fromkeys(
            s(inv.get("shipping_mark")) for inv in invoices if inv.get("shipping_mark")
        ))
 
        MARK_START = 38
        ROWS_PER   = 5
        MAX_INV    = 100
 
        # ── helper: สีพื้นหลัง ──────────────────────────────────
        HIGHLIGHT_FILL = PatternFill(fill_type="solid", fgColor="BDD7EE")
 
        # ── คำนวณ row ก่อนเพื่อให้รู้ bl_row ──────────────────────
        cargo_end = MARK_START + len(invoices) * ROWS_PER
        hs_row    = cargo_end + 2
        fr_row    = hs_row + 2
        bl_row    = fr_row + 1
 
        # ════════════════════════════════════════════════════════
        # STEP 1: ล้าง border ทั้งหมด row 34 → bl_row+20
        # ════════════════════════════════════════════════════════
        for r in range(34, bl_row + 20):
            for c in range(1, 12):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell, MergedCell): continue
                cell.border = Border()
 
        # ════════════════════════════════════════════════════════
        # STEP 2: เคลียร์ค่า (value) ในพื้นที่ cargo
        # ════════════════════════════════════════════════════════
        clear_end = MARK_START + MAX_INV * ROWS_PER + 2
        for r in range(MARK_START, clear_end):
            for c in range(1, 12):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell, MergedCell): continue
                if isinstance(cell.value, str) and cell.value.startswith("="): continue
                cell.value = None
 
        # ════════════════════════════════════════════════════════
        # STEP 3: วาด border ใหม่ทีละ column (row34 → bl_row)
        # ════════════════════════════════════════════════════════
        thin = Side(border_style="thin")
 
        # col A (1): left+right ทุก row, bottom เฉพาะ bl_row
        for r in range(34, bl_row + 1):
            cell = ws.cell(row=r, column=1)
            if isinstance(cell, MergedCell): continue
            cell.border = Border(
                left=thin, right=thin,
                bottom=thin if r == bl_row else Side(border_style=None)
            )
 
        # col B (2): left ทุก row, bottom เฉพาะ bl_row
        for r in range(34, bl_row + 1):
            cell = ws.cell(row=r, column=2)
            if isinstance(cell, MergedCell): continue
            cell.border = Border(
                left=thin,
                bottom=thin if r == bl_row else Side(border_style=None)
            )
 
        # col C (3): left ทุก row, bottom เฉพาะ bl_row
        for r in range(34, bl_row + 1):
            cell = ws.cell(row=r, column=3)
            if isinstance(cell, MergedCell): continue
            cell.border = Border(
                left=thin,
                bottom=thin if r == bl_row else Side(border_style=None)
            )
 
        # col D-G (4-7): ไม่มี left/right (พื้นที่ description), bottom เฉพาะ bl_row
        for r in range(34, bl_row + 1):
            for c in range(4, 8):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell, MergedCell): continue
                cell.border = Border(
                    bottom=thin if r == bl_row else Side(border_style=None)
                )
 
        # col H (8): right ทุก row, bottom เฉพาะ bl_row
        for r in range(34, bl_row + 1):
            cell = ws.cell(row=r, column=8)
            if isinstance(cell, MergedCell): continue
            cell.border = Border(
                right=thin,
                bottom=thin if r == bl_row else Side(border_style=None)
            )
 
        # col I (9): right ทุก row, bottom เฉพาะ bl_row
        for r in range(34, bl_row + 1):
            cell = ws.cell(row=r, column=9)
            if isinstance(cell, MergedCell): continue
            cell.border = Border(
                right=thin,
                bottom=thin if r == bl_row else Side(border_style=None)
            )
 
        # col J (10): left+right ทุก row, bottom เฉพาะ bl_row
        for r in range(34, bl_row + 1):
            cell = ws.cell(row=r, column=10)
            if isinstance(cell, MergedCell): continue
            cell.border = Border(
                left=thin, right=thin,
                bottom=thin if r == bl_row else Side(border_style=None)
            )
 
        # ════════════════════════════════════════════════════════
        # STEP 4: เพิ่ม top border row 34 (ใต้ header row 33)
        # ════════════════════════════════════════════════════════
        for c in range(1, 11):
            cell = ws.cell(row=34, column=c)
            if isinstance(cell, MergedCell): continue
            old = cell.border
            cell.border = Border(top=thin, left=old.left, right=old.right, bottom=old.bottom)
 
        # Shipping Marks ลง col A — แยกแต่ละบรรทัดลงคนละ row
        mark_row = MARK_START
        for i, mark in enumerate(marks[:MAX_INV]):
            if i > 0:
                mark_row += 1  # blank row between invoices
            for line in str(mark).split("\n"):
                line = line.strip()
                if line:
                    safe_write_rc(mark_row, 1, line)
                    mark_row += 1
 
        def _parse_pkg(val):
            m = re.match(r'^([\d,]+)\s+(.*)', str(val or '').strip())
            if m:
                try:
                    num  = int(m.group(1).replace(',', ''))
                    unit = m.group(2).strip()
                    if unit.upper().startswith("PP."):
                        unit = unit[3:]
                    return num, unit
                except Exception:
                    pass
            return 0, "CARTONS"

        first_pkg_unit = _parse_pkg(invoices[0].get("cartons"))[1] if invoices else "CARTONS"

        gw_cells, cbm_cells, ctn_cells = [], [], []
 
        for idx, inv in enumerate(invoices):
            base = MARK_START + idx * ROWS_PER
            for offset in range(ROWS_PER):
                copy_row_style(38 + offset, base + offset)
 
            # col B = "CARTONS" เฉพาะ invoice แรกเท่านั้น
            if idx == 0:
                safe_write_rc(base, 2, first_pkg_unit)
 
            # row+0: qty — ใส่สีเฉพาะ C (จำนวน + unit เช่น "1,389 CARTONS" หรือ "40 PP.PALLETS")
            qty_num, pkg_unit = _parse_pkg(inv.get("cartons"))
            safe_write_rc(base, 3, qty_num or 0)
            safe_write_rc(base, 4, pkg_unit)
            safe_write_rc(base, 5, f"({s(inv.get('quantity_str'))})")
            cell_c = ws.cell(row=base, column=3)
            if not isinstance(cell_c, MergedCell):
                cell_c.fill = copy.copy(HIGHLIGHT_FILL)
 
            # row+1: description
            safe_write_rc(base+1, 3, s(inv.get("description")))
 
            # row+2: invoice no.
            safe_write_rc(base+2, 3, f"INVOICE NO. {s(inv.get('invoice_no'))}")
 
            # row+3: G.W. — ใส่สีเฉพาะ D (ตัวเลข GW) และ G (ตัวเลข CBM)
            safe_write_rc(base+3, 3, "G.W.  ")
            safe_write_rc(base+3, 4, inv.get("gross_weight_kgs") or 0)
            safe_write_rc(base+3, 5, "KGS")
            safe_write_rc(base+3, 6, "M3")
            safe_write_rc(base+3, 7, inv.get("measurement_cbm") or 0)
            safe_write_rc(base+3, 8, "CBM")
            for col in [4, 7]:   # D=ตัวเลข GW, G=ตัวเลข CBM
                cell = ws.cell(row=base+3, column=col)
                if not isinstance(cell, MergedCell):
                    cell.fill = copy.copy(HIGHLIGHT_FILL)
 
            gw_cells.append(f"D{base+3}")
            cbm_cells.append(f"G{base+3}")
            ctn_cells.append(f"C{base}")
 
        if gw_cells:
            safe_write("I35", "=" + "+".join(gw_cells))
            safe_write("J35", "=" + "+".join(cbm_cells))
        if ctn_cells:
            # B37 = SUM ตัวเลข C ของทุก invoice
            safe_write("B37", "=" + "+".join(ctn_cells))
 
        # ── HS Code, Freight, BL ─────────────────────────────────
        # bl_row / hs_row / fr_row คำนวณและวาด border ไว้แล้วใน STEP 1-4
 
        for r in [hs_row, fr_row, bl_row]:
            try:
                ws.merge_cells(f"C{r}:H{r}")
            except Exception:
                pass
            # ล้าง right border ที่ merge_cells สร้างให้อัตโนมัติ
            for c in range(3, 9):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell, MergedCell): continue
                old = cell.border
                cell.border = Border(
                    top=old.top, left=old.left,
                    right=Side(border_style=None),
                    bottom=old.bottom,
                )
 
        hs = s(extra.get("hs_code_all"))
        if hs:
            safe_write_rc(hs_row, 3, f"HS CODE : {hs}")
            cell = ws.cell(row=hs_row, column=3)
            if not isinstance(cell, MergedCell):
                cell.alignment = Alignment(horizontal="center")
 
        safe_write_rc(fr_row, 3, s(extra.get("freight_terms")) or "FREIGHT COLLECT")
        safe_write_rc(bl_row, 3, s(extra.get("bl_type")) or "Sea Waybill")
        for r in [fr_row, bl_row]:
            cell = ws.cell(row=r, column=3)
            if not isinstance(cell, MergedCell):
                cell.alignment = Alignment(horizontal="center")
 
        # ── Container table (optional — แสดงเฉพาะเมื่อ user กรอก cont_no) ──
        filled_containers = [c for c in containers if s(c.get("cont_no"))]
 
        if filled_containers:
            CONT_START = bl_row + 2
            ws.row_dimensions[CONT_START].height = ws.row_dimensions[59].height
            for c, hdr in enumerate(
                ["CONT. NO.","SEAL NO.","QTY","TYPE OF PACKAGE","GW.","M3","SIZE CONT","TARE WEIGHT","DT","VGM","HS CODE"],
                start=1
            ):
                safe_write_rc(CONT_START, c, hdr)
 
            for idx, cont in enumerate(filled_containers):
                r = CONT_START + 1 + idx
                ws.row_dimensions[r].height = ws.row_dimensions[60].height
                safe_write_rc(r, 1,  s(cont.get("cont_no")))
                safe_write_rc(r, 2,  s(cont.get("seal_no")))
                safe_write_rc(r, 3,  cont.get("cartons"))
                safe_write_rc(r, 4,  "CARTONS")
                safe_write_rc(r, 5,  cont.get("gw"))
                safe_write_rc(r, 6,  cont.get("cbm"))
                safe_write_rc(r, 7,  s(cont.get("size")) or "40 ' HQ")
                safe_write_rc(r, 8,  cont.get("tare"))
                safe_write_rc(r, 9,  cont.get("dt") or 10)
                safe_write_rc(r, 10, f"=E{r}+H{r}+I{r}")
                safe_write_rc(r, 11, s(cont.get("hs_code")))
 
            last_cont  = CONT_START + len(filled_containers)
            total_cont = last_cont + 1
            copy_row_style(65, total_cont)
            safe_write_rc(total_cont, 3, f"=SUM(C{CONT_START+1}:C{last_cont})")
            safe_write_rc(total_cont, 5, f"=SUM(E{CONT_START+1}:E{last_cont})")
            safe_write_rc(total_cont, 6, f"=SUM(F{CONT_START+1}:F{last_cont})")
 
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
 
    # ── UI ──────────────────────────────────────────────────────
    st.markdown("""
    <div style="display:flex;align-items:center;gap:16px;padding:14px 20px;
                background:#ffffff;border:1px solid #e8e8e8;border-left:4px solid #D40511;
                border-radius:12px;margin-bottom:20px;box-shadow:0 2px 8px rgba(0,0,0,0.05);">
        <div style="background:#FFCC00;padding:8px 16px;border-radius:6px;">
            <span style="color:#D40511;font-family:'Arial Black',sans-serif;font-size:22px;font-weight:900;">DHL</span>
        </div>
        <div>
            <div style="color:#111;font-weight:700;font-size:17px;">Auto SI Generator</div>
            <div style="color:#999;font-size:12px;">เลือก Booking → อัปโหลด Invoice → Generate SI.xlsx</div>
        </div>
    </div>
    """, unsafe_allow_html=True)
 
    # Step 1: Booking
    st.markdown("**Step 1 · เลือก Booking**")
    booking_map = {}
    try:
        res2 = supabase.table(TBL_BOOKINGS).select("*").order("updated_at", desc=True).execute()
        if res2.data:
            for row in res2.data:
                if row.get("booking_no"):
                    booking_map[row["booking_no"]] = row
    except Exception as e:
        st.error(f"Load bookings error: {e}")
 
    bk_options  = ["-- เลือก Booking No. --"] + list(booking_map.keys())
    selected_bk = st.selectbox("Booking No.", bk_options, label_visibility="collapsed")
    bk          = booking_map.get(selected_bk, {})
 
    if bk:
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Booking No.", bk.get("booking_no","—"))
        m2.metric("Container", f"{bk.get('no_container','—')}×{bk.get('container_type','—')}")
        m3.metric("Paperless", bk.get("paperless_code","—"))
        m4.metric("SI Cutoff", bk.get("si_cutoff","—"))
 
    st.divider()
 
    # Step 2: Invoice PDFs
    st.markdown("**Step 2 · อัปโหลด Invoice PDF** (โยนพร้อมกันหลายใบได้)")
    inv_files = st.file_uploader(
        "Invoice PDFs", type="pdf", accept_multiple_files=True, key="si_inv_up"
    )
 
    st.divider()
 
    # Step 3: SI Template
    st.markdown("**Step 3 · อัปโหลด SI Template (.xlsx)**")
    tmpl_file = st.file_uploader("SI Template", type="xlsx", key="si_tmpl_up")
 
    st.divider()
 
    # Step 4: ข้อมูลเสริม
    st.markdown("**Step 4 · ข้อมูลเสริม**")
    col_e1, col_e2 = st.columns(2)
    with col_e1:
        freight_terms = st.selectbox("Freight Terms", ["FREIGHT COLLECT","FREIGHT PREPAID"])
        bl_type       = st.selectbox("BL Type", ["Sea Waybill","Original B/L","Surrender B/L","Telex Release"])
        revised       = st.checkbox("REVISED", value=False)
    with col_e2:
        hs_code_input = st.text_input("HS Code (คั่นด้วย , )", placeholder="8415.10, 3926.90")
 
    # Container table
    st.markdown("**ข้อมูล Container**")
    n_cont = st.number_input("จำนวน Container", min_value=1, max_value=20,
                              value=int(bk.get("no_container") or 1))
 
    cont_header = st.columns([2,2,1.2,1.2,1.2,1.2,1.2,1])
    for h, col in zip(["CONT. NO.","SEAL NO.","CARTONS","G.W.(KGS)","CBM","TARE(KGS)","SIZE","DT"], cont_header):
        col.markdown(f"<div style='font-size:11px;font-weight:600;color:#999;'>{h}</div>", unsafe_allow_html=True)
 
    container_rows = []
    for i in range(int(n_cont)):
        cols = st.columns([2,2,1.2,1.2,1.2,1.2,1.2,1])
        container_rows.append({
            "cont_no": cols[0].text_input("", key=f"cno_{i}",  placeholder=f"CONT {i+1}", label_visibility="collapsed"),
            "seal_no": cols[1].text_input("", key=f"sno_{i}",  placeholder="SEAL",        label_visibility="collapsed"),
            "cartons": cols[2].number_input("", key=f"ctn_{i}", min_value=0, value=0,      label_visibility="collapsed"),
            "gw":      cols[3].number_input("", key=f"cgw_{i}", min_value=0.0, value=0.0, label_visibility="collapsed", format="%.2f"),
            "cbm":     cols[4].number_input("", key=f"ccb_{i}", min_value=0.0, value=0.0, label_visibility="collapsed", format="%.3f"),
            "tare":    cols[5].number_input("", key=f"ctr_{i}", min_value=0, value=3900,   label_visibility="collapsed"),
            "size":    cols[6].selectbox("",   key=f"csz_{i}", options=["40 ' HQ","40 ' GP","20 ' GP"], label_visibility="collapsed"),
            "dt":      cols[7].number_input("", key=f"cdt_{i}", min_value=0, value=10,    label_visibility="collapsed"),
            "hs_code": hs_code_input,
        })
 
    st.divider()
 
    can_gen = selected_bk != "-- เลือก Booking No. --" and inv_files and tmpl_file
    if not can_gen:
        st.info("⬆️ กรุณาเลือก Booking + อัปโหลด Invoice PDF + SI Template ก่อน")
 
    if can_gen and st.button("🚀 GENERATE SI.xlsx", use_container_width=True):
        all_invoices = []
        prog = st.progress(0)
        for idx, f in enumerate(inv_files):
            with st.spinner(f"กำลังอ่าน: {f.name}"):
                try:
                    extracted = _extract_invoices(f.read())
                    all_invoices.extend(extracted)
                    st.success(f"✅ {f.name} → {len(extracted)} invoice(s)")
                except Exception as e:
                    st.error(f"❌ {f.name}: {e}")
            prog.progress((idx+1)/len(inv_files))
 
        if not all_invoices:
            st.error("ไม่พบข้อมูล Invoice")
        else:
            # Preview
            st.markdown("---")
            st.markdown("**📋 Preview — ตรวจสอบก่อน Generate**")
            st.dataframe(pd.DataFrame([{
                "Invoice No.": inv.get("invoice_no"),
                "Description": inv.get("description"),
                "Cartons":     inv.get("cartons"),
                "QTY":         inv.get("quantity_str"),
                "GW (KGS)":    inv.get("gross_weight_kgs"),
                "CBM":         inv.get("measurement_cbm"),
                "Vessel":      inv.get("vessel_feeder"),
                "Port Disc.":  inv.get("port_of_discharge"),
                "ETD":         inv.get("etd"),
            } for inv in all_invoices]), use_container_width=True)
 
            # Generate
            with st.spinner("กำลังสร้างไฟล์ SI..."):
                try:
                    si_bytes = _fill_si(
                        template_bytes = tmpl_file.read(),
                        booking        = bk,
                        invoices       = all_invoices,
                        containers     = container_rows,
                        extra          = {
                            "hs_code_all":   hs_code_input,
                            "freight_terms": freight_terms,
                            "bl_type":       bl_type,
                            "revised":       revised,
                        },
                    )
                    inv_nos  = "_".join(inv.get("invoice_no","") for inv in all_invoices)
                    filename = f"SI_{bk.get('booking_no','BK')}_INV_{inv_nos}.xlsx"
                    st.success("🎉 สร้างไฟล์ SI สำเร็จ!")
                    st.download_button(
                        "📥 ดาวน์โหลด SI.xlsx",
                        data=si_bytes, file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                except Exception as e:
                    st.error(f"❌ Generate Error: {e}")
                    import traceback; st.code(traceback.format_exc())

# ─────────────────────────────────────────
# 9. PAGE: LOCAL CHARGES
# ─────────────────────────────────────────
if page == "💰 Local Charges":

    st.subheader("💰 Local Charges")

    st.markdown("""
    <style>
    button[data-testid="stNumberInputStepUp"],
    button[data-testid="stNumberInputStepDown"] { display: none; }
    </style>
    """, unsafe_allow_html=True)

    # ── Booking No. dropdown ──────────────
    try:
        bk_res = supabase.table(TBL_BOOKINGS).select("booking_no").order("updated_at", desc=True).execute()
        bk_options = [r["booking_no"] for r in bk_res.data if r.get("booking_no")]
    except Exception:
        bk_options = []

    selected_booking_no = st.selectbox(
        "Booking No. *",
        options=["— เลือก Booking —"] + bk_options,
    )
    booking_selected = selected_booking_no != "— เลือก Booking —"
    if not booking_selected:
        st.warning("⚠️ กรุณาเลือก Booking No. ก่อนอัปโหลดไฟล์")

    # ── Fetch shipment type for DSV WHT rule ──
    shipment_type = ""
    if booking_selected:
        try:
            bk_detail = supabase.table(TBL_BOOKINGS).select("by_air_or_sea, fcl_or_lcl").eq("booking_no", selected_booking_no).limit(1).execute()
            if bk_detail.data:
                b = bk_detail.data[0]
                air_sea = (b.get("by_air_or_sea") or "").strip().lower()
                fcl_lcl = (b.get("fcl_or_lcl") or "").strip().lower()
                if air_sea == "air":
                    shipment_type = "Air Export"
                elif fcl_lcl == "lcl":
                    shipment_type = "Ocean Export"
                else:
                    shipment_type = "Ocean Export"
        except Exception:
            pass

    # ── Upload zone ──────────────────────
    if "lc_uploader_key" not in st.session_state:
        st.session_state["lc_uploader_key"] = 0

    lc_file = st.file_uploader(
        "โยนไฟล์ Local Charge PDF ที่นี่",
        type="pdf",
        accept_multiple_files=False,
        key=f"lc_{st.session_state['lc_uploader_key']}",
        disabled=not booking_selected,
    )

    if lc_file and booking_selected:
        cache_key = f"lc_data_{lc_file.name}_{lc_file.size}"
        if cache_key not in st.session_state:
            with st.spinner(f"กำลังสแกน: {lc_file.name}"):
                try:
                    st.session_state[cache_key] = extract_local_charges(lc_file.read(), shipment_type=shipment_type)
                except Exception as e:
                    st.error(f"❌ AI Error: {e}")
                    st.session_state[cache_key] = None

        data = st.session_state[cache_key]

        if data:
            if data.get("_multi_invoice"):
                st.warning("⚠️ ตรวจพบหลาย invoice ในไฟล์เดียวกัน — แสดงเฉพาะ invoice แรก กรุณาอัพโหลดทีละ invoice")
            st.success("✅ Extract สำเร็จ — ตรวจสอบข้อมูลก่อนบันทึก")

            # ── Header fields ─────────────────────────────────────
            st.markdown("**ข้อมูลทั่วไป**")
            r1c1, r1c2, r1c3 = st.columns(3)
            hdr_agent_invoice_no = r1c1.text_input("Agent Invoice No.", value=str(data.get("agent_invoice_no") or ""), key="lc_agent_invoice_no")
            hdr_pay_to           = r1c2.text_input("Pay To",            value=str(data.get("pay_to")        or ""), key="lc_pay_to")
            hdr_tax_id           = r1c3.text_input("Tax ID No.",        value=str(data.get("tax_id")        or ""), key="lc_tax_id")

            r2c1, r2c2, r2c3 = st.columns(3)
            hdr_tax_name      = r2c1.text_area("Tax Name & Address", value=str(data.get("tax_name")      or ""), height=80, key="lc_tax_name")
            hdr_delivery_port = r2c2.text_input("Delivery Port",     value=str(data.get("delivery_port") or ""), key="lc_delivery_port")
            hdr_etd           = r2c3.text_input("ETD (DD/MM/YYYY)",  value=str(data.get("etd")           or ""), key="lc_etd")

            r3c1, r3c2, r3c3 = st.columns(3)
            hdr_bl_no         = r3c1.text_input("B/L No.",         value=str(data.get("bl_no") or ""), key="lc_bl_no")
            hdr_ctc_invoice_no = r3c2.text_input("CTC Invoice No.", value="", key="lc_ctc_invoice_no")
            hdr_remark        = r3c3.text_input("Remark",           value="", key="lc_remark")

            r4c1, r4c2, r4c3 = st.columns(3)
            hdr_due_date      = r4c1.text_input("Due Date (DD/MM/YYYY)", value=str(data.get("due_date") or ""), key="lc_due_date")

            # ── Dynamic items ─────────────────────────────────────
            st.markdown("**ค่าใช้จ่าย (บาท)**")

            CATEGORY_ORDER = [
                "thc_40hc", "thc_40dv", "thc_20gp",
                "export_handling", "seal", "bl_fee", "surrender_fee", "vgm_fee",
                "doc_amendment", "detention", "demurrage", "container_repair",
                "edi_fee", "late_gate", "environmental_fee", "storage", "freight_charge", "other",
            ]
            CATEGORY_LABEL = {
                "thc_40hc":          "THC (40HC)",
                "thc_40dv":          "THC (40DV)",
                "thc_20gp":          "THC (20GP)",
                "export_handling":   "Export Handling",
                "seal":              "Seal",
                "bl_fee":            "B/L Fee",
                "surrender_fee":     "Surrender Fee",
                "vgm_fee":           "VGM Coordination Fee",
                "doc_amendment":     "Documentation Amendment Charge",
                "detention":         "Detention",
                "demurrage":         "Demurrage",
                "container_repair":  "Container Repair",
                "edi_fee":           "EDI Fee",
                "late_gate":         "Late Gate Service",
                "environmental_fee": "Environmental Fee",
                "storage":           "Storage",
                "freight_charge":    "Freight Charge",
                "other":             None,  # keep original description
            }

            if "lc_items" not in st.session_state or st.session_state.get("lc_items_source") != cache_key:
                raw_items = [
                    {"description": it.get("description", ""), "category": it.get("category") or "other",
                     "wht_pct": int(it.get("wht_pct") or 0),
                     "rate": float(it.get("rate") or 0), "qty": float(it.get("qty") or 0),
                     "total": float(it.get("total") or 0)}
                    for it in (data.get("items") or [])
                ]
                # Replace description with standard label (keep original for "other")
                for it in raw_items:
                    cat = it["category"]
                    label = CATEGORY_LABEL.get(cat)
                    if label:
                        it["description"] = label
                # Sort by category order
                raw_items.sort(key=lambda x: CATEGORY_ORDER.index(x["category"]) if x["category"] in CATEGORY_ORDER else 99)
                st.session_state["lc_items"] = raw_items
                st.session_state["lc_items_source"] = cache_key

            hc1, hc2, hc3, hc4, hc5, hc6 = st.columns([3, 1, 2, 2, 2, 1])
            hc1.markdown("**รายการ**"); hc2.markdown("**WHT %**")
            hc3.markdown("**Rate**"); hc4.markdown("**No.Unit**")
            hc5.markdown("**Total**"); hc6.markdown("")

            items_to_delete = []
            for idx, item in enumerate(st.session_state["lc_items"]):
                c1, c2, c3, c4, c5, c6 = st.columns([3, 1, 2, 2, 2, 1])
                item["description"] = c1.text_input("_", value=item["description"], label_visibility="collapsed", key=f"lc_desc_{idx}")
                item["wht_pct"]     = int(c2.number_input("_", value=int(item["wht_pct"]), min_value=0, max_value=3, step=1, label_visibility="collapsed", key=f"lc_wht_{idx}"))
                item["rate"]        = c3.number_input("_", value=item["rate"], min_value=0.0, step=0.01, format="%.2f", label_visibility="collapsed", key=f"lc_rate_{idx}")
                item["qty"]         = c4.number_input("_", value=item["qty"],  min_value=0.0, step=0.01, format="%.2f", label_visibility="collapsed", key=f"lc_qty_{idx}")
                item["total"]       = c5.number_input("_", value=item["total"], min_value=0.0, step=0.01, format="%.2f", label_visibility="collapsed", key=f"lc_total_{idx}")
                if c6.button("🗑️", key=f"lc_del_{idx}"):
                    items_to_delete.append(idx)

            for idx in reversed(items_to_delete):
                st.session_state["lc_items"].pop(idx)
                st.rerun()

            if st.button("➕ เพิ่มรายการ", key="lc_add"):
                st.session_state["lc_items"].append({"description": "", "category": "other", "wht_pct": 0, "rate": 0.0, "qty": 0.0, "total": 0.0})
                st.rerun()

            # ── Live summary ──────────────────────────────────────
            current_items = st.session_state["lc_items"]
            charges_subtotal = sum(float(it.get("total") or 0) for it in current_items)
            wht1_sum = sum(float(it.get("total") or 0) for it in current_items if int(it.get("wht_pct") or 0) == 1)
            wht3_sum = sum(float(it.get("total") or 0) for it in current_items if int(it.get("wht_pct") or 0) == 3)
            calc_wht1 = round(wht1_sum * 0.01, 2)
            calc_wht3 = round(wht3_sum * 0.03, 2)

            st.divider()
            sub1, _, _, _, sub5, _ = st.columns([3, 1, 2, 2, 2, 1])
            sub1.markdown("<div style='padding-top:8px'>**รวมค่าใช้จ่าย**</div>", unsafe_allow_html=True)
            sub5.markdown(f"<div style='padding-top:8px; text-align:right'><b>{charges_subtotal:,.2f}</b></div>", unsafe_allow_html=True)

            st.markdown("**สรุป**")
            sc1, _, _, _, sc5, _ = st.columns([3, 1, 2, 2, 2, 1])
            sc1.markdown("<div style='padding-top:8px'>VAT 7%</div>", unsafe_allow_html=True)
            vat_7 = sc5.number_input("_", value=float(data.get("vat_7") or 0), min_value=0.0, step=0.01, format="%.2f", label_visibility="collapsed", key="vat_7")

            after_vat = charges_subtotal + vat_7
            av1, _, _, _, av5, _ = st.columns([3, 1, 2, 2, 2, 1])
            av1.markdown("<div style='padding-top:8px'>**รวมหลัง VAT**</div>", unsafe_allow_html=True)
            av5.markdown(f"<div style='padding-top:8px; text-align:right'><b>{after_vat:,.2f}</b></div>", unsafe_allow_html=True)

            wc1, _, _, _, wc5, _ = st.columns([3, 1, 2, 2, 2, 1])
            wc1.markdown("<div style='padding-top:8px'>WHT 1%</div>", unsafe_allow_html=True)
            wc5.markdown(f"<div style='padding-top:8px; text-align:right'>{calc_wht1:,.2f}</div>", unsafe_allow_html=True)

            w3c1, _, _, _, w3c5, _ = st.columns([3, 1, 2, 2, 2, 1])
            w3c1.markdown("<div style='padding-top:8px'>WHT 3%</div>", unsafe_allow_html=True)
            w3c5.markdown(f"<div style='padding-top:8px; text-align:right'>{calc_wht3:,.2f}</div>", unsafe_allow_html=True)

            after_wht = round(after_vat - calc_wht1 - calc_wht3, 2)
            aw1, _, _, _, aw5, _ = st.columns([3, 1, 2, 2, 2, 1])
            aw1.markdown("<div style='padding-top:8px'>**รวมหลังหัก WHT**</div>", unsafe_allow_html=True)
            aw5.markdown(f"<div style='padding-top:8px; text-align:right'><b>{after_wht:,.2f}</b></div>", unsafe_allow_html=True)

            # ── Save ──────────────────────────────────────────────
            if st.button("💾 บันทึก", use_container_width=True, key="lc_save"):
                header = {
                    "agent_invoice_no": hdr_agent_invoice_no or None,
                    "pay_to":        hdr_pay_to or None,
                    "tax_name":      hdr_tax_name or None,
                    "tax_id":        hdr_tax_id or None,
                    "delivery_port": hdr_delivery_port or None,
                    "etd":           hdr_etd or None,
                    "bl_no":         hdr_bl_no or None,
                    "due_date":      hdr_due_date or None,
                    "vat_7":         vat_7 if vat_7 else None,
                    "wht_1":         calc_wht1 if calc_wht1 else None,
                    "wht_3":         calc_wht3 if calc_wht3 else None,
                    "subtotal":      charges_subtotal if charges_subtotal else None,
                    "total":         after_wht if after_wht else None,
                    "source_file":   lc_file.name,
                    "booking_no":    selected_booking_no if selected_booking_no != "— เลือก Booking —" else None,
                    "ctc_invoice_no": hdr_ctc_invoice_no or None,
                    "remark":        hdr_remark or None,
                }
                save_items = [
                    {"description": it["description"], "category": it.get("category") or "other",
                     "wht_pct": it["wht_pct"],
                     "rate": it["rate"] or None, "qty": it["qty"] or None,
                     "total": it["total"] or None}
                    for it in current_items if it.get("description")
                ]
                if save_local_charge_v2(header, save_items, pdf_bytes=lc_file.getvalue(), filename=lc_file.name):
                    st.success("✅ บันทึกเรียบร้อยแล้ว")
                    del st.session_state["lc_items"]
                    del st.session_state["lc_items_source"]
                    st.session_state["lc_uploader_key"] += 1
                    st.rerun()

    # ── History table ────────────────────
    st.divider()
    st.subheader("📊 รายการ Local Charges ทั้งหมด")
    try:
        lc_res = (
            supabase.table(TBL_LOCAL_CHARGES_V2)
            .select("*")
            .order("created_at", desc=True)
            .execute()
        )
        if lc_res.data:
            df_lc = pd.DataFrame(lc_res.data)
            df_lc = bkk_time(df_lc, "created_at")
            st.dataframe(df_lc, use_container_width=True)

            st.markdown("**ลบรายการ**")
            delete_options = {
                f"{r.get('agent_invoice_no') or r.get('ctc_invoice_no') or '-'} | {r.get('booking_no') or '-'} | {r.get('pay_to') or '-'} | {r.get('etd') or '-'}": r["id"]
                for r in lc_res.data
            }
            selected_label = st.selectbox("เลือกรายการที่ต้องการลบ", options=["— เลือก —"] + list(delete_options.keys()), key="lc_delete_select")
            if selected_label != "— เลือก —":
                if st.button("🗑️ ลบรายการนี้", type="primary", key="lc_delete_btn"):
                    del_id = delete_options[selected_label]
                    try:
                        supabase.table(TBL_LOCAL_CHARGE_ITEMS).delete().eq("local_charge_id", del_id).execute()
                        supabase.table(TBL_LOCAL_CHARGES_V2).delete().eq("id", del_id).execute()
                        st.success("✅ ลบเรียบร้อยแล้ว")
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ ลบไม่สำเร็จ: {e}")
        else:
            st.info("ยังไม่มีข้อมูล")
    except Exception as e:
        st.error(f"❌ Load Error: {e}")

# ─────────────────────────────────────────
# 10. PAGE: EXPORT SUMMARY
# ─────────────────────────────────────────
if page == "📊 Export Summary":
    st.subheader("📊 Export Summary")

    # ── Load all booking numbers that have local charges ──
    try:
        from collections import defaultdict
        lc_res = supabase.table(TBL_LOCAL_CHARGES_V2).select("id,booking_no,ctc_invoice_no,exported_at").execute()
        bno_rows = defaultdict(list)
        for r in lc_res.data:
            if r.get("booking_no"):
                bno_rows[r["booking_no"]].append(r)
    except Exception as e:
        st.error(f"❌ Load Error: {e}")
        bno_rows = {}

    if not bno_rows:
        st.info("ยังไม่มีข้อมูล Local Charges ในระบบ")
    else:
        def _make_label(bno, rows):
            all_exported = all(r.get("exported_at") for r in rows)
            icon = "✅" if all_exported else "⚠️"
            ctc_list = ", ".join(filter(None, (r.get("ctc_invoice_no") for r in rows))) or "—"
            return f"{icon} {bno}  [{ctc_list}]"

        label_to_bno = {_make_label(bno, rows): bno for bno, rows in bno_rows.items()}
        sorted_labels = sorted(label_to_bno)

        selected_labels = st.multiselect(
            "เลือก Booking No.",
            options=sorted_labels,
            placeholder="เลือกได้หลาย Booking No.",
        )
        selected_bnos = [label_to_bno[l] for l in selected_labels]

        # เคลียร์ PDF cache เมื่อ selection เปลี่ยน (คงสถานะ checkbox ไว้)
        if st.session_state.get("_export_bnos") != selected_bnos:
            st.session_state.pop("export_pdf", None)
            st.session_state["_export_bnos"] = selected_bnos
        if "export_checks" not in st.session_state:
            st.session_state["export_checks"] = {}

        if selected_bnos:
            try:
                # ── โหลดข้อมูลทั้งหมดสำหรับ preview ──
                all_records = []   # [{header, items}, ...]
                preview_rows = []  # rows สำหรับ DataFrame

                # โหลด no_container / no_pallet / country จาก bookings table
                bk_res = (
                    supabase.table(TBL_BOOKINGS)
                    .select("booking_no,no_container,no_pallet,country")
                    .in_("booking_no", selected_bnos)
                    .execute()
                )
                bk_map = {r["booking_no"]: r for r in (bk_res.data or [])}

                for bno in selected_bnos:
                    hdrs = (
                        supabase.table(TBL_LOCAL_CHARGES_V2)
                        .select("*")
                        .eq("booking_no", bno)
                        .execute()
                    ).data
                    for hdr in hdrs:
                        its = (
                            supabase.table(TBL_LOCAL_CHARGE_ITEMS)
                            .select("*")
                            .eq("local_charge_id", hdr["id"])
                            .execute()
                        ).data
                        subtotal = sum(float(it.get("total") or 0) for it in its)
                        vat_7 = float(hdr.get("vat_7") or 0)
                        wht_1 = float(hdr.get("wht_1") or 0)
                        wht_3 = float(hdr.get("wht_3") or 0)
                        net   = subtotal + vat_7 - wht_1 - wht_3
                        bk    = bk_map.get(bno, {})
                        all_records.append({"header": hdr, "items": its, "bk": bk})
                        preview_rows.append({
                            "เลือก":          st.session_state["export_checks"].get(hdr["id"], True),
                            "Booking No.":    bno,
                            "CTC Invoice":    hdr.get("ctc_invoice_no") or "—",
                            "Pay To":         hdr.get("pay_to") or "—",
                            "ยอดสุทธิ (THB)": round(net, 2),
                            "Delivery Port":  hdr.get("delivery_port") or "—",
                            "No. Container":  bk.get("no_container") or "—",
                            "No. Pallet":     bk.get("no_pallet") or "—",
                            "สถานะ":          "✅ Exported" if hdr.get("exported_at") else "⚠️ ยังไม่ export",
                        })

                # ── ตารางตัวอย่าง + เลือก/ไม่เลือก ──
                st.markdown("**ตัวอย่างข้อมูล** — ติ๊กถูกรายการที่ต้องการ Export")
                df_preview = pd.DataFrame(preview_rows)
                edited = st.data_editor(
                    df_preview,
                    column_config={
                        "เลือก":          st.column_config.CheckboxColumn("เลือก", default=True),
                        "ยอดสุทธิ (THB)": st.column_config.NumberColumn("ยอดสุทธิ (THB)", format="%.2f"),
                    },
                    disabled=["Booking No.", "CTC Invoice", "Pay To", "ยอดสุทธิ (THB)", "Delivery Port", "No. Container", "No. Pallet", "สถานะ"],
                    hide_index=True,
                    use_container_width=True,
                    key="export_preview_editor",
                )

                # บันทึกสถานะ checkbox กลับใน session state
                for i, rec in enumerate(all_records):
                    st.session_state["export_checks"][rec["header"]["id"]] = bool(edited.iloc[i]["เลือก"])

                n_selected = int(edited["เลือก"].sum())
                st.caption(f"เลือกอยู่ {n_selected} / {len(edited)} รายการ")

                # ── Cover Page Preview & Editor ──────────────────────
                st.divider()
                st.markdown("**ตัวอย่าง Cover Page** — แก้ไข รายการ / Remark ได้ก่อน Export")
                cover_rows = []
                for i, rec in enumerate(all_records):
                    hdr = rec["header"]
                    bk  = rec.get("bk") or {}
                    its = rec["items"]
                    subtotal = sum(float(it.get("total") or 0) for it in its)
                    vat_7 = float(hdr.get("vat_7") or 0)
                    wht_1 = float(hdr.get("wht_1") or 0)
                    wht_3 = float(hdr.get("wht_3") or 0)
                    net   = round(subtotal + vat_7 - wht_1 - wht_3, 2)
                    cats  = sorted(set(it.get("category") or "other" for it in its))
                    default_part = "+ ".join(c.upper().replace("_", "") for c in cats)
                    cover_rows.append({
                        "_id":         hdr["id"],
                        "รายการ":      default_part,
                        "Invoice No.": hdr.get("ctc_invoice_no") or "-",
                        "Country":     bk.get("country") or "-",
                        "Pay To":      hdr.get("pay_to") or "-",
                        "Due Date":    hdr.get("due_date") or "-",
                        "Remark":      hdr.get("remark") or "",
                        "Amount":      net,
                    })

                df_cover = pd.DataFrame(cover_rows)
                edited_cover = st.data_editor(
                    df_cover,
                    column_config={
                        "_id":    st.column_config.Column("_id", disabled=True),
                        "Amount": st.column_config.NumberColumn("Amount (THB)", format="%.2f", disabled=True),
                    },
                    column_order=["รายการ", "Invoice No.", "Country", "Pay To", "Due Date", "Remark", "Amount"],
                    disabled=["_id", "Amount"],
                    hide_index=True,
                    use_container_width=True,
                    key="cover_page_editor",
                )

                # ฉีด cover fields เข้าใน all_records (ใช้เฉพาะ cover page PDF)
                cover_edit_map = {row["_id"]: row for _, row in edited_cover.iterrows()}
                for rec in all_records:
                    eid = rec["header"]["id"]
                    if eid in cover_edit_map:
                        rec["cover_part"]    = cover_edit_map[eid]["รายการ"]
                        rec["cover_inv"]     = cover_edit_map[eid]["Invoice No."]
                        rec["cover_country"] = cover_edit_map[eid]["Country"]
                        rec["cover_payto"]   = cover_edit_map[eid]["Pay To"]
                        rec["cover_due"]     = cover_edit_map[eid]["Due Date"]
                        rec["cover_remark"]  = cover_edit_map[eid]["Remark"]

                st.divider()
                st.markdown("**ผู้รับผิดชอบ**")
                _PREPARED_BY = {
                    "Sutida Suwantatree":       "098-584-2550",
                    "Nattawan Kerdpol":         "094-914-0449",
                    "Alyssa Daenglang":         "095-363-0707",
                    "Waraphon Praneetpolkrang": "062-396-4024",
                    "Wannisa Seesuksam":        "066-109-7538",
                    "Nisachol Pongmulee":       "083-110-3758",
                }
                _pc1, _pc2 = st.columns(2)
                prepared_name  = _pc1.selectbox("ชื่อผู้รับผิดชอบ", options=["— เลือก —"] + list(_PREPARED_BY), key="export_prepared_name")
                prepared_phone = _PREPARED_BY.get(prepared_name, "")
                _pc2.markdown("**เบอร์โทรศัพท์**")
                _pc2.write(prepared_phone if prepared_phone else "—")
                if prepared_name == "— เลือก —":
                    prepared_name = ""

                if st.button("📄 Generate PDF", use_container_width=True, disabled=(n_selected == 0)):
                    filtered = [all_records[i] for i, chk in enumerate(edited["เลือก"]) if chk]
                    if filtered:
                        st.session_state["export_pdf"]      = generate_expense_pdf(filtered, prepared_by=prepared_name, prepared_by_phone=prepared_phone)
                        st.session_state["export_filename"] = "expense_summary_" + "_".join(selected_bnos) + ".pdf"
                        # Mark exported rows in Supabase
                        from datetime import datetime, timezone
                        now_iso = datetime.now(timezone.utc).isoformat()
                        exported_ids = [all_records[i]["header"]["id"] for i, chk in enumerate(edited["เลือก"]) if chk]
                        for eid in exported_ids:
                            supabase.table(TBL_LOCAL_CHARGES_V2).update({"exported_at": now_iso}).eq("id", eid).execute()
                    else:
                        st.warning("กรุณาเลือกอย่างน้อย 1 รายการ")

            except Exception as e:
                st.error(f"❌ Load Error: {e}")

        if st.session_state.get("export_pdf"):
            st.download_button(
                label="⬇️ ดาวน์โหลด PDF",
                data=st.session_state["export_pdf"],
                file_name=st.session_state.get("export_filename", "export.pdf"),
                mime="application/pdf",
                use_container_width=True,
            )
